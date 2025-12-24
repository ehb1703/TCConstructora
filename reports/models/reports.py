# -*- coding: utf-8 -*-
from odoo import api, fields, models
from odoo.exceptions import UserError, ValidationError
import logging
import os
import tempfile
import openpyxl
import binascii
from datetime import datetime
from odoo.tools import DEFAULT_SERVER_DATE_FORMAT

_logger = logging.getLogger(__name__)


class reportWeeks(models.Model):
    _name = 'report.weeks'
    _inherit = ['mail.thread', 'mail.activity.mixin']
    _description = 'Semanas'
    _rec_name = 'no_semana'

    no_semana = fields.Integer(string='Número de Semana')
    anio = fields.Integer(string='Año de la semana')
    finicio = fields.Date(string='Fecha de inicio')
    ffin = fields.Date(string='Fecha de termino')


class reportWeeks(models.Model):
    _name = 'report.empresas.obras'
    _inherit = ['mail.thread', 'mail.activity.mixin']
    _description = 'Empresas - Obras'
    _rec_name = 'obra'
    
    empresa = fields.Char(string='Empresa')
    obra = fields.Char(string='Obra')


class reportRequisition(models.Model):
    _name = 'report.requisition'
    _inherit = ['mail.thread', 'mail.activity.mixin']
    _description = 'Informe de Requisiciones'
    _rec_name = 'name'
    _order = 'year, week_id'

    week_id = fields.Many2one('report.weeks', string='No. de Semana')
    year = fields.Integer(string='Año')
    name = fields.Char('Informe')
    file = fields.Binary(string='Archivo', help='Seleccionar el archivo con el formato correcto para la carga la información.')
    filename = fields.Char(string='Nombre del archivo', tracking=True)
    line_ids = fields.One2many(comodel_name='report.requisition.line', inverse_name='req_id', string='Lineas')
    data_ids = fields.One2many(comodel_name='report.requisition.data', inverse_name='req_id', string='Lineas a cargar')
    state = fields.Selection(selection=[('draft', 'Borrador'), ('done', 'Hecho'), ('duplicado', 'Duplicado')], 
        string = 'Estado', default = 'draft')
    act_informacion = fields.Boolean(string='Cargar información de empresas - obras', default=False)

    def datos_cabecera(self):
        self.env.cr.execute("""SELECT semana, t3.anio_periodo, (t3.anio||'-'||mi||'-'||dia_inicial)::date inicio, (t3.anio_periodo||'-'||mf||'-'||dia_final)::date fin
            FROM (SELECT t2.semana, t2.anio, t2.anio_periodo, t2.dia_inicial, t2.dia_final,
                        (CASE WHEN SUBSTRING(t2.mes_inicial, 1, 3) = 'ENE' THEN '01' WHEN SUBSTRING(t2.mes_inicial, 1, 3) = 'FEB' THEN '01' 
                            WHEN SUBSTRING(t2.mes_inicial, 1, 3) = 'MAR' THEN '03' WHEN SUBSTRING(t2.mes_inicial, 1, 3) = 'ABR' THEN '04' 
                            WHEN SUBSTRING(t2.mes_inicial, 1, 3) = 'MAY' THEN '05' WHEN SUBSTRING(t2.mes_inicial, 1, 3) = 'JUN' THEN '06' 
                            WHEN SUBSTRING(t2.mes_inicial, 1, 3) = 'JUL' THEN '07' WHEN SUBSTRING(t2.mes_inicial, 1, 3) = 'AGO' THEN '08' 
                            WHEN SUBSTRING(t2.mes_inicial, 1, 3) = 'SEP' THEN '09' WHEN SUBSTRING(t2.mes_inicial, 1, 3) = 'OCT' THEN '10' 
                            WHEN SUBSTRING(t2.mes_inicial, 1, 3) = 'NOVIEMBRE' THEN '11' ELSE '12' END) mi,
                        (CASE WHEN SUBSTRING(t2.mes_final, 1, 3) = 'ENE' THEN '01' WHEN SUBSTRING(t2.mes_final, 1, 3) = 'FEB' THEN '01' 
                            WHEN SUBSTRING(t2.mes_final, 1, 3) = 'MAR' THEN '03' WHEN SUBSTRING(t2.mes_final, 1, 3) = 'ABR' THEN '04' 
                            WHEN SUBSTRING(t2.mes_final, 1, 3) = 'MAY' THEN '05' WHEN SUBSTRING(t2.mes_final, 1, 3) = 'JUN' THEN '06' 
                            WHEN SUBSTRING(t2.mes_final, 1, 3) = 'JUL' THEN '07' WHEN SUBSTRING(t2.mes_final, 1, 3) = 'AGO' THEN '08' 
                            WHEN SUBSTRING(t2.mes_final, 1, 3) = 'SEP' THEN '09' WHEN SUBSTRING(t2.mes_final, 1, 3) = 'OCT' THEN '10' 
                            WHEN SUBSTRING(t2.mes_final, 1, 3) = 'NOV' THEN '11' ELSE '12' END) mf
                FROM (SELECT t1.periodo, t1.semana, substring(t1.periodo, length(t1.periodo) - 3, 4) anio_periodo,
                        (CASE WHEN t1.espacios in (8, 9) THEN substring(t1.periodo, length(t1.periodo) - 3, 4)
                            ELSE substring(t1.periodo, POSITION('AL' IN t1.periodo) - 5, 4) end) anio,
                        (CASE  WHEN t1.espacios = 8 THEN SPLIT_PART(t1.PERIODO, ' ', 5) WHEN t1.espacios >= 9 THEN SPLIT_PART(t1.PERIODO, ' ', 3) 
                            ELSE 'REVISAR EL CASO' END) mes_inicial,
                        (CASE WHEN t1.espacios = 8 THEN SPLIT_PART(t1.PERIODO, ' ', 5) WHEN t1.espacios = 9 THEN SPLIT_PART(t1.PERIODO, ' ', 6) 
                            ELSE SPLIT_PART(t1.PERIODO, ' ', 7) END) mes_final,
                        TRIM(substring(t1.periodo, POSITION('DEL' IN t1.periodo) + 4, 2)) dia_inicial, 
                        TRIM(substring(t1.periodo, POSITION('AL' IN t1.periodo) + 3, 2)) dia_final
                    FROM (SELECT regexp_count('""" + self.filename + "', ' ') espacios, SPLIT_PART(SPLIT_PART('" + self.filename + 
            "', ' ', 3), '.', 1) semana, TRIM(SPLIT_PART('" + self.filename + "', '.', 2)) periodo) as t1) as t2) as t3 ")
        week = self.env.cr.dictfetchall()

        self.env.cr.execute('SELECT MIN(ID) id, COUNT(*) num FROM report_weeks WHERE no_semana = ' + str(week[0]['semana']) + ' AND anio = ' 
            + str(week[0]['anio_periodo']) + " AND finicio = '" + str(week[0]['inicio']) + "' AND ffin = '" + str(week[0]['fin']) + "'")
        existe = self.env.cr.dictfetchall()

        if existe[0]['num'] == 0:
            semana_dict = {'no_semana': week[0]['semana'], 'anio': week[0]['anio_periodo'], 'finicio': week[0]['inicio'], 'ffin': week[0]['fin']}
            week_id = self.env['report.weeks'].create(semana_dict)
        else:
            week_id = self.env['report.weeks'].search([('id', '=', existe[0]['id'])])

        return week_id


    def validarDatos(self):
        self.env.cr.execute("""SELECT count(*) num
            FROM (SELECT rrd.col2, reo.empresa FROM report_requisition_data rrd LEFT JOIN report_empresas_obras reo ON rrd.col2 = reo.obra 
                WHERE rrd.req_id = """ + str(self.id) + " AND rrd.col3 = 'DESCRIPCION' GROUP BY 1, 2) AS t1 WHERE t1.empresa IS NULL")
        empresas = self.env.cr.dictfetchall()

        if empresas[0]['num'] != 0:
            self.act_informacion = True
        else:
            self.act_informacion = False


    def action_cargar_registros(self):
        for record in self:
            if not record.file:
                raise ValidationError('Seleccione un archivo para cargar.')

            if record.file and record.data_ids:
                raise ValidationError('Ya hay información cargada. En caso de ser necesario volver a cargar debe eliminarlos.')

            filename, file_extension = os.path.splitext(record.filename)
            if file_extension in ['.xlsx', '.xls', '.xlsm']:
                record.__leer_carga_archivo()
            else:
                raise ValidationError('Seleccione un archivo tipo xlsx, xls.')


    def __leer_carga_archivo(self):
        for record in self:
            week = record.datos_cabecera()

            file = tempfile.NamedTemporaryFile(suffix=".xlsx")
            file.write(binascii.a2b_base64(record.file))
            file.seek(0)
            xlsx_file = file.name
            
            wb = openpyxl.load_workbook(filename=xlsx_file, data_only=True)
            sheets = wb.sheetnames
            sheet_name = sheets[0]
            sheet = wb[sheet_name]
            registros = []
            for row in sheet.iter_rows(values_only=True):
                col1 = str(row[0]).strip()
                col2 = str(row[1]).strip()
                col3 = str(row[2]).strip()
                col4 = str(row[3]).strip()
                col5 = str(row[4]).strip()
                col6 = str(row[5]).strip()
                col7 = str(row[6]).strip()
                col8 = str(row[7]).strip()
                if col1 == 'None':
                    col1 = ''
                if col2 == 'None':
                    col2 = ''
                if col3 == 'None':
                    col3 = ''
                if col4 == 'None':
                    col4 = ''
                if col5 == 'None':
                    col5 = ''
                if col6 == 'None':
                    col6 = ''
                if col7 == 'None':
                    col7 = ''
                if col8 == 'None':
                    col8 = ''
                registro = {'col1': col1, 'col2': col2, 'col3': col3, 'col4': col4, 'col5': col5, 'col6': col6, 'col7': col7, 'col8': col8}
                registros.append((0, 0, registro))

            existe = self.env['report.requisition'].search([('week_id', '=', week.id)])
            if existe:
                state = 'duplicado'
            else:
                state = 'draft'

            record.write({'data_ids': registros, 'state': state, 'week_id': week.id, 'year': week.anio, 'name': str(week.no_semana) + ' - ' + str(week.anio)})

    def action_genera_req(self):
        for record in self:
            info = record.validarDatos()
            if record.act_informacion:
                raise ValidationError('No existe relación entre la empresa y la obra. Favor de vincularla')

            s = 0
            registros = []
            empresa = ''
            obra = ''            
            for data in record.data_ids:
                if data.col1 == data.col2 == data.col3 == data.col4 == data.col5 == data.col6 == data.col7 == data.col8 == '':
                    s += 1
                else:
                    if data.col7 == '' or data.col1[:5].upper() == 'TOTAL' or data.col2[:5].upper() == 'TOTAL' or data.col5[:5].upper() == 'TOTAL' \
                            or data.col1 in ('EMPRESA', 'REQUISICION SEMANAL'):
                        s += 1
                    else:
                        if data.col1 != '':
                            empresa = data.col1
                        if data.col2 != '':
                            obra = data.col2
                        if data.col5 == '-':
                            fuerza = 0
                        else:
                            fuerza = data.col5
                        if data.col3 == 'DESCRIPCION':
                            empresa = self.env['report.empresas.obras'].search([('obra', '=', obra)]).empresa
                        elif data.col3 == '':
                            _logger.warning('No se agrega el renglón')
                        else:
                            if data.col8 == '':
                                adeudo = 0
                            else:
                                adeudo = float(data.col8)
                            registro = {'empresa': empresa, 'obra': obra, 'concepto': data.col3, 'proveedor': data.col4, 'fuerza': fuerza, 
                                'metodo_pago': data.col6, 'importe': data.col7, 'adeudo': adeudo}
                            registros.append((0, 0, registro))
            record.write({'line_ids': registros, 'state': 'done'})

    def action_vincular(self):
        self.env.cr.execute('''INSERT INTO report_empresas_obras (CREATE_DATE, WRITE_DATE, CREATE_UID, WRITE_UID, OBRA) 
            SELECT NOW(), NOW(), ''' + str(self.env.user.id) + ', ' + str(self.env.user.id) + ', rrd.col2 FROM report_requisition_data rrd WHERE rrd.req_id = ' + 
            str(self.id) + ''' AND rrd.col3 = 'DESCRIPCION' AND NOT EXISTS (SELECT * FROM report_empresas_obras reo WHERE rrd.col2 = reo.obra) 
            GROUP BY 1, 2, 3, 4, 5 ''')

        view_tree = self.env.ref('reports.view_report_empresas_obras_list')
        view_form = self.env.ref('reports.view_reports_empresas_obras_form')
        self.act_informacion = False
        
        action = {
            'name': ('Empresas - Obras'),
            'view_type': 'form',
            'view_mode': 'list, form',
            'res_model': 'report.empresas.obras',
            'views': [(view_tree.id, 'list'),(view_form.id, 'form')],
            'view_id': view_tree.id,
            'type': 'ir.actions.act_window',
            'domain': [('empresa', '=', None)] }
        return action

    def action_unlink_details(self):
        for record in self:
            record.data_ids.unlink()


class reportRequisitionLine(models.Model):
    _name = 'report.requisition.line'
    _description = 'Lineas del Informe de Requisiciones'
    
    req_id = fields.Many2one(comodel_name='report.requisition', string='Requisición', readonly=True)
    empresa = fields.Char(string='Empresa')
    obra = fields.Char(string='Obra')
    concepto = fields.Char(string='Concepto')
    proveedor = fields.Char(string='Proveedor')
    fuerza = fields.Integer(string='Fuerza')
    metodo_pago = fields.Char(string='Forma de Pago')
    importe = fields.Float(string='Importe')
    adeudo = fields.Float(string='Adeudo')


class reportRequisitionData(models.Model):
    _name = 'report.requisition.data'
    _description = 'Información del archivo de la requisición semanal'
    
    req_id = fields.Many2one(comodel_name='report.requisition', string='Requisición', readonly=True)
    col1 = fields.Char(string='Columna 1')
    col2 = fields.Char(string='Columna 2')
    col3 = fields.Char(string='Columna 3')
    col4 = fields.Char(string='Columna 4')
    col5 = fields.Char(string='Columna 5')
    col6 = fields.Char(string='Columna 6')
    col7 = fields.Char(string='Columna 7')
    col8 = fields.Char(string='Columna 8')
