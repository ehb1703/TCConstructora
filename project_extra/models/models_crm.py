# -*- coding: utf-8 -*-
from odoo import api, fields, models, _
from odoo.exceptions import UserError, ValidationError
from odoo.tools import html_escape
from markupsafe import Markup
from dateutil.relativedelta import relativedelta
import logging
import os
import tempfile
import openpyxl
import binascii
import base64
import io

_logger = logging.getLogger(__name__)

class CrmRevertLog(models.Model):
    _name = 'crm.revert.log'
    _description = 'Bitácora de reversiones de etapa'

    lead_id = fields.Many2one('crm.lead', string='Oportunidad', required=True, ondelete='cascade')
    user_id = fields.Many2one('res.users', string='Usuario', default=lambda self: self.env.user, required=True)
    old_stage_id = fields.Many2one('crm.stage', string='Etapa anterior', required=True)
    new_stage_id = fields.Many2one('crm.stage', string='Etapa nueva', required=True)
    reason_id = fields.Many2one('crm.revert.reason', string='Motivo')
    reason_text = fields.Text(string='Observaciones')

class CrmLead(models.Model):
    _inherit = 'crm.lead'

    zona_geografica_id = fields.Many2one('project.zona.geografica', string='Zona geográfica', tracking=True)
    partner_emisor_id = fields.Many2one('res.partner', string='Dependencia emisora', tracking=True)
    tipo_obra_id = fields.Many2one('project.type', string='Tipo de obra', tracking=True)
    especialidad_ids = fields.Many2many('project.especialidad', string='Especialidad(es) requerida(s)')
    monto_min = fields.Float(string='Monto mínimo')
    monto_max = fields.Float(string='Monto máximo')
    fecha_convocatoria = fields.Date(string='Fecha de convocatoria')
    fecha_limite_inscripcion = fields.Date(string='Fecha límite de inscripción')
    fecha_apertura = fields.Date(string='Fecha de apertura')
    convocatoria_pdf = fields.Binary(string='PDF de convocatoria', attachment=True)
    convocatoria_pdf_name = fields.Char(string='Nombre del archivo')
    origen_id = fields.Many2one('crm.lead.type', string='Tipo de Venta')
    origen_name = fields.Char(string='Tipo nombre', compute='_compute_bases')
    req_bases = fields.Boolean(string='Requiere pago de bases', compute='_compute_bases')
    tipo_obra_ok = fields.Boolean('Tipo de obra cumple', tracking=True)
    dependencia_ok = fields.Boolean('Dependencia emisora cumple', tracking=True)
    capital_ok = fields.Boolean('Capital contable cumple', tracking=True)
    in_calificado = fields.Boolean(string='En calificado', default=False)
    oc_ids = fields.One2many('purchase.order', 'lead_id', string='Ordenes de compra relacionada')
    revert_log_ids = fields.One2many('crm.revert.log', 'lead_id', string='Bitácora de reversiones', readonly=True)
    revert_log_count = fields.Integer(compute='_compute_revert_log_count', string='Reversiones')
    no_licitacion = fields.Char(string='No. de Procedimiento')
    desc_licitacion = fields.Char(string='Descripción')
    currency_id = fields.Many2one('res.currency', string='Moneda', tracking=True)
    stage_name = fields.Char(string='State name', compute='_compute_name_stage', store=False)
    stage_previous = fields.Char(string='Etapa anterior')
    # Inscripción / Compra de bases
    bases_pay = fields.Boolean('Pagar bases', tracking=True)
    bases_pagado = fields.Boolean('Pagado', readonly=True, tracking=True, help='Se marca automáticamente cuando la factura de bases está pagada')
    bases_supervisor_id = fields.Many2one('hr.employee', string='Supervisor general', tracking=True)
    bases_cost = fields.Float(string='Costo de las bases', tracking=True)
    bases_doc = fields.Binary(string='Docto. Bases', attachment=True)
    bases_doc_name = fields.Char(string='Nombre del documento')
    bases_notification_sent = fields.Boolean(string='Notificación de bases enviada', default=False)
    bases_anticipo_porcentaje = fields.Integer(string='% de anticipo', tracking=True)
    bases_abstinencia_anticipo = fields.Boolean(string='Abstinencia de anticipo', tracking=True)
    bases_modalidad_contrato_id = fields.Many2one('project.modalidad.contrato', string='Modalidad de contrato', tracking=True)
    bases_fecha_inicio_trabajos = fields.Date(string='Fecha estimada para inicio de trabajos', tracking=True)
    bases_fecha_terminacion_trabajos = fields.Date(string='Fecha estimada para terminación de trabajos', tracking=True)
    bases_plazo_ejecucion = fields.Integer(string='Plazo de ejecución', tracking=True, help='Plazo de ejecución en días naturales')
    bases_sancion_atraso = fields.Boolean(string='Sanción por atraso', tracking=True, help='Si no se cumple el tiempo estipulado se generan sanciones')
    bases_ret_5_millar = fields.Boolean(string='Ret. 5 al millar', tracking=True)
    bases_ret_2_millar = fields.Boolean(string='Ret. 2 al millar', tracking=True)
    ptdcto_ids = fields.Many2many('project.docsrequeridos', 'propuesta_tecnica_doctos_rel', 'tecnica_id', 'docto_id', string='Documentos requeridos',
        domain="[('model_id', '=', 'crm.lead'), ('etapa','=','tecnica')]")
    pedcto_ids = fields.Many2many('project.docsrequeridos', 'propuesta_economica_doctos_rel', 'economica_id', 'docto_id', string='Documentos requeridos',
        domain="[('model_id', '=', 'crm.lead'), ('etapa','=','economica')]")
    # Visita de obra
    visita_obligatoria = fields.Boolean(string='Asistencia obligatoria')
    visita_personas_ids = fields.Many2many('hr.employee', 'crm_lead_visita_employee_rel', 'lead_id', 'employee_id', string='Personas asignadas')
    visita_fecha = fields.Datetime(string='Fecha y hora de visita')
    visita_lugar_reunion = fields.Char(string='Lugar de reunión', size=200, help='Dirección y ubicación de la visita de obra')
    visita_acta = fields.Binary(string='Acta de visita', attachment=True)
    visita_acta_name = fields.Char(string='Nombre acta de visita')
    visita_notif_auto_sent = fields.Boolean(string='Notif. Automática enviada', default=False)
    visita_notif_manual_sent = fields.Boolean(string='Notif. Manual enviada', default=False)
    # Junta de Aclaración de dudas
    junta_obligatoria = fields.Boolean(string='Asistencia obligatoria')
    junta_notif_auto_sent = fields.Boolean(string='Notif. Automática enviada', default=False)
    junta_notif_manual_sent = fields.Boolean(string='Notif. Manual enviada', default=False)
    junta_line_ids = fields.One2many('crm.junta.line', 'lead_id', string='Juntas de aclaración de dudas')
    circular_ids = fields.One2many('crm.circular.line', 'lead_id', string='Circulares')
    # Asignación de responsable
    tecnico_documental_id = fields.Many2one('hr.employee', 'Técnico/documental', tracking=True)
    analista_id = fields.Many2one('crm.analyst', 'Analista', tracking=True)
    economico_operativo_id = fields.Many2one('hr.employee', 'Económico/operativo', tracking=True)
    junta_dudas_notif_auto_sent = fields.Boolean(string='Notif. Automática fecha límite dudas enviada', default=False)
    # Conceptos de obra
    doctoconcept_id = fields.Many2one('report.tipodocumento', string='Tipo de documento Catálogo de Conceptos', 
        domain="[('module_id', '=', 'crm.lead'), ('no_docto', '=', '2')]")
    doctomatriz_id = fields.Many2one('report.tipodocumento', string='Tipo de documento Matriz', 
        domain="[('module_id', '=', 'crm.lead'), ('no_docto', '=', '9')]")
    doctobasicos_id = fields.Many2one('report.tipodocumento', string='Tipo de documento Básicos', 
        domain="[('module_id', '=', 'crm.lead'), ('no_docto', '=', '10')]")
    concept_ids = fields.One2many('crm.concept.line', 'lead_id', string='Conceptos de trabajo')
    budget_ids = fields.One2many('crm.budget.line', 'lead_id', string='Partidas presupestales')
    combo_ids = fields.One2many('crm.combo.line', 'lead_id', string='Combos')
    basico_ids = fields.One2many('crm.basico.line', 'lead_id', string='Básicos')
    relacion_ids = fields.One2many('crm.basico.relacion', 'lead_id', string='Relación básicos')
    # Insumos
    input_ids = fields.One2many('crm.input.line', 'lead_id', string='Insumos')
    input_file = fields.Binary(string='Archivo', attachment=True)
    input_filename = fields.Char(string='Nombre del archivo', tracking=True)
    no_cotizar_insumos = fields.Boolean(string='No se necesita cotizar insumos', default=False)
    # Junta de Apertura de Propuestas
    apertura_obligatoria = fields.Boolean('Asistencia obligatoria')
    apertura_personas_ids = fields.Many2many('hr.employee', 'crm_lead_apertura_employee_rel', 'lead_id', 'employee_id', 'Personas asignadas')
    apertura_fecha = fields.Datetime('Fecha y hora de junta')
    apertura_fecha_limite = fields.Date('Fecha límite')
    apertura_lugar_reunion = fields.Char(string='Lugar de reunión', size=200, help='Dirección y ubicación de la junta de apertura de propuestas')
    apertura_acta = fields.Binary('Acta de la junta', attachment=True)
    apertura_acta_name = fields.Char('Nombre acta de la junta')
    apertura_notif_auto_sent = fields.Boolean('Notif. Automática enviada', default=False)
    apertura_notif_manual_sent = fields.Boolean('Notif. Manual enviada', default=False)
    # Junta de Fallo
    fallo_obligatoria = fields.Boolean(string='Asistencia obligatoria')
    fallo_personas_ids = fields.Many2many('hr.employee', 'crm_lead_fallo_employee_rel', 'lead_id', 'employee_id', string='Personas asignadas')
    fallo_fecha = fields.Datetime(string='Fecha y hora de junta')
    fallo_lugar_reunion = fields.Char(string='Lugar de reunión', size=200, help='Dirección y ubicación de la junta de pronunciamiento del fallo')
    fallo_fecha_adjudicacion = fields.Date(string='Fecha de adjudicación')
    fallo_ganado = fields.Boolean(string='Ganado')
    fallo_notif_auto_sent = fields.Boolean(string='Notif. Automática enviada', default=False)
    fallo_notif_manual_sent = fields.Boolean(string='Notif. Manual enviada', default=False)
    fallo_notif_directores_auto_sent = fields.Boolean(string='Notif. Automática dirección enviada', default=False)
    fallo_notif_directores_manual_sent = fields.Boolean(string='Notif. Manual dirección enviada', default=False)
    fallo_acta = fields.Binary(string='Acta de fallo', attachment=True)
    fallo_acta_name = fields.Char(string='Nombre acta de fallo')
    # Ganado
    fecha_limite_firma = fields.Date('Fecha límite de firma')
    fecha_firma = fields.Date('Fecha de firma')
    contrato_firmado = fields.Boolean('Contrato firmado')
    contrato_documento = fields.Binary('Contrato', attachment=True)
    contrato_documento_name = fields.Char('Nombre del contrato')
    importe_contratado = fields.Float('Importe contratado', tracking=True)
    importe_anticipo = fields.Float('Importe anticipo', compute='_compute_importe_anticipo', store=True, readonly=True)
    rupc_siop = fields.Char('RUPC. SIOP', size=50)
    no_estimaciones = fields.Integer('No. De estimaciones')
    es_siop = fields.Boolean('Es SIOP', compute='_compute_es_siop', store=False)
    orden_trabajo_doc = fields.Binary('Orden de trabajo', attachment=True, help='Documento de orden de trabajo entregado por la dependencia')
    orden_trabajo_doc_name = fields.Char('Nombre orden de trabajo')
    # Nuevos campos etapa Ganado
    empresa_concursante_id = fields.Many2one('res.company', string='Empresa concursante', help='Empresa del grupo que participa en la licitación')
    clave_unidad_sat_id = fields.Many2one('product.unspsc.code', string='Clave Unidad de Medida',
        domain="[('applies_to', '=', 'uom')]")
    clave_producto_servicio_id = fields.Many2one('product.unspsc.code', string='Clave producto/Servicio',
        domain="[('segmento_sat', '=', '72'), ('applies_to', '=', 'product')]")
    fianza_anticipo_no = fields.Char(string='Fianza anticipo No.', size=100, help='Número de fianza del anticipo')
    afianzadora_anticipo_id = fields.Many2one('res.partner', string='Nombre afianzadora', domain="[('is_afianzadora', '=', True), ('is_company', '=', True)]")
    fianza_cumplimiento_no = fields.Char(string='Fianza cumplimiento No.', size=100, help='Número de fianza de cumplimiento')
    afianzadora_cumplimiento_id = fields.Many2one('res.partner', string='Nombre afianzadora cumplimiento', 
        domain="[('is_afianzadora', '=', True), ('is_company', '=', True)]")
    director_estimaciones_pagos = fields.Char(string='Director(a) de estimaciones y pagos', size=100, 
        help='Nombre del director responsable de estimaciones y pagos')
    #Propuesta Técnica y Económica
    documents_folder_id = fields.Many2one('documents.document', string="Folder", copy=False,domain="[('type', '=', 'folder'), ('shortcut_document_id', '=', False), '|', ('company_id', '=', False), ('company_id', '=', company_id)]",)
    documents_count = fields.Integer('Documentos', compute='_compute_documents_count', readonly=True)
    documents_count_tecnica = fields.Integer('Docs Técnica', compute='_compute_documents_count', readonly=True)
    documents_count_economica = fields.Integer('Docs Económica', compute='_compute_documents_count', readonly=True)
    pt_doc_line_ids = fields.One2many('crm.propuesta.tecnica.doc', 'lead_id', string='Documentos Propuesta Técnica')
    pt_revision_ids = fields.One2many('crm.propuesta.tecnica.revision', 'lead_id', string='Revisiones Propuesta Técnica')
    pe_doc_line_ids = fields.One2many('crm.propuesta.economica.doc', 'lead_id', string='Documentos Propuesta Económica')
    pe_revision_ids = fields.One2many('crm.propuesta.economica.revision', 'lead_id', string='Revisiones Propuesta Económica')
    pe_autorizado = fields.Boolean(string='PE Autorizada', default=False)
    #Campos de Autorización de Presupuesto
    pe_presupuesto_subtotal = fields.Float(string='Subtotal', readonly=True, digits=(16, 2))
    pe_presupuesto_iva = fields.Float(string='IVA', readonly=True, digits=(16, 2))
    pe_presupuesto_total = fields.Float(string='Total', readonly=True, digits=(16, 2))
    pe_presupuesto_revisado = fields.Boolean(string='Revisado', default=False)
    pe_presupuesto_autorizado = fields.Boolean(string='Autorizado', default=False)

    @api.depends('importe_contratado', 'bases_anticipo_porcentaje', 'bases_abstinencia_anticipo')
    def _compute_importe_anticipo(self):
        for lead in self:
            if lead.bases_abstinencia_anticipo or not lead.bases_anticipo_porcentaje:
                lead.importe_anticipo = 0.0
            else:
                lead.importe_anticipo = lead.importe_contratado * (lead.bases_anticipo_porcentaje / 100.0)

    @api.depends('no_licitacion')
    def _compute_es_siop(self):
        for lead in self:
            if lead.no_licitacion and len(lead.no_licitacion) >= 4:
                lead.es_siop = lead.no_licitacion[:4].upper() == 'SIOP'
            else:
                lead.es_siop = False

    @api.onchange('origen_id')
    def _compute_bases(self):
        for record in self:
            record.req_bases = bool(getattr(record.origen_id, 'bases', False))
            record.origen_name = record.origen_id.name if record.origen_id else False

    @api.depends('revert_log_ids')
    def _compute_revert_log_count(self):
        groups = self.env['crm.revert.log'].read_group([('lead_id', 'in', self.ids)], ['lead_id'], ['lead_id'])
        counts = {g['lead_id'][0]: g['lead_id_count'] for g in groups}
        for r in self:
            r.revert_log_count = counts.get(r.id, 0)

    def _compute_name_stage(self):
        for record in self:
            record.stage_name = record.stage_id.name

    @api.depends('pe_doc_line_ids', 'pt_doc_line_ids')
    def _compute_documents_count(self):
        Document = self.env['documents.document']
        for lead in self:
            try:
                if 'documents.document' not in self.env:
                    lead.documents_count = 0
                    lead.documents_count_tecnica = 0
                    lead.documents_count_economica = 0
                    continue
                
                lic_name = lead.no_licitacion or lead.name or 'Sin nombre'
                root = Document.search([('name', '=', 'CRM'), ('type', '=', 'folder'), ('folder_id', '=', False)], limit=1)
                if not root:
                    lead.documents_count = 0
                    lead.documents_count_tecnica = 0
                    lead.documents_count_economica = 0
                    continue
                
                lic_folder = Document.search([('name', '=', lic_name), ('type', '=', 'folder'), ('folder_id', '=', root.id)], limit=1)
                if not lic_folder:
                    lead.documents_count = 0
                    lead.documents_count_tecnica = 0
                    lead.documents_count_economica = 0
                    continue
                
                tecnica_folder = Document.search([('name', '=', 'Tecnico'), ('type', '=', 'folder'), ('folder_id', '=', lic_folder.id)], limit=1)
                economica_folder = Document.search([('name', '=', 'Economico'), ('type', '=', 'folder'), ('folder_id', '=', lic_folder.id)], limit=1)
                
                # Contar documentos SEPARADOS por carpeta
                count_tecnica = 0
                count_economica = 0
                
                if tecnica_folder:
                    count_tecnica = Document.search_count([('folder_id', '=', tecnica_folder.id), ('type', '!=', 'folder')])
                if economica_folder:
                    count_economica = Document.search_count([('folder_id', '=', economica_folder.id), ('type', '!=', 'folder')])
                
                lead.documents_count_tecnica = count_tecnica
                lead.documents_count_economica = count_economica
                lead.documents_count = count_tecnica + count_economica
                
            except Exception as e:
                _logger.error("Error contando documentos para lead %s: %s" % (lead.id, str(e)))
                lead.documents_count = 0
                lead.documents_count_tecnica = 0
                lead.documents_count_economica = 0


    # ---------- Helpers ----------
    def _get_stage_by_name(self, name):
        # Busca etapa por nombre EXACTO en el pipeline del lead (o global).
        self.ensure_one()
        Stage = self.env['crm.stage']
        domain = [('name', '=', name)]
        if self.team_id:
            domain = ['|', ('team_id', '=', self.team_id.id), ('team_id', '=', False)] + domain
        return Stage.search(domain, limit=1)

    def _ensure_stage_is_fallo(self):
        # Asegura que la etapa actual sea 'Junta de Fallo' (por nombre).
        self.ensure_one()
        if self.env.context.get('allow_lost_any_stage'):
            return

        stage_name = (self.stage_id.name or '').strip().lower()
        if stage_name != 'junta de fallo':
            raise UserError(_('Solo puede marcar Perdido en la etapa JUNTA DE FALLO.'))


    def _get_authorizer_emails_from_group(self, grupo):
        # Obtiene correos de los usuarios del grupo project_extra.group_conv_authorizer.
        emails = set()
        group = self.env.ref(grupo, raise_if_not_found=False)
        if group:
            for user in group.users.filtered(lambda u: u.active and u.partner_id and u.partner_id.email):
                emails.add(user.partner_id.email.strip())
        return sorted(e for e in emails if '@' in e)

    def _post_html(self, title, old_stage=None, new_stage=None):
        parts = [f'<p>{html_escape(title)}</p>']
        if old_stage or new_stage:
            parts.append(
                f'<p>{html_escape(_('De'))} <b>{html_escape((old_stage and old_stage.name) or '-')}</b> '
                f'{html_escape(_('a'))} <b>{html_escape((new_stage and new_stage.name) or '-')}</b>.</p>')
        body = '<div>' + ''.join(parts) + '</div>'
        self.message_post(body=Markup(body), message_type='comment', subtype_xmlid='mail.mt_note')

    def _log_stage_change(self, old_stage, new_stage, reason_id, reason_text=''):
        self.env['crm.revert.log'].sudo().create({'lead_id': self.id, 'user_id': self.env.user.id, 'old_stage_id': old_stage.id if old_stage else False,
            'new_stage_id': new_stage.id if new_stage else False, 'reason_id': reason_id, 'reason_text': reason_text or False,})

    def _sync_pe_doc_lines(self):
        # Sincroniza la tabla de Propuesta Económica con los documentos seleccionados en pedcto_ids.
        DocLine = self.env['crm.propuesta.economica.doc']
        for lead in self:
            docs = lead.pedcto_ids
            existing_by_doc = {line.docto_id.id: line for line in lead.pe_doc_line_ids}
            wanted_ids = set(docs.ids)

            # Crear líneas nuevas para documentos que faltan
            for doc in docs:
                if doc.id not in existing_by_doc:
                    DocLine.create({'lead_id': lead.id, 'docto_id': doc.id})

            # Eliminar líneas de documentos que ya no están seleccionados
            for line in lead.pe_doc_line_ids:
                if line.docto_id.id not in wanted_ids:
                    line.unlink()


    @api.onchange('pedcto_ids')
    def _onchange_pedcto_ids(self):
        #Cuando cambian los documentos requeridos de propuesta económica, actualizamos la tabla.
        if self.id:
            self._sync_pe_doc_lines()

    def check_doctos(self, type):
        # Verifica que haya al menos 2 personas diferentes que hayan revisado los documentos.
        for lead in self:
            if type == 'tecnica':
                revisadas = lead.pt_doc_line_ids.filtered(lambda r: not r.generado)
            else:
                revisadas = lead.pe_doc_line_ids.filtered(lambda r: not r.generado)
            
            if len(revisadas) != 0:
                raise UserError(_('Existen archivos sin cargar'))


    def check_revisions(self, type):
        # Verifica que haya al menos 2 personas diferentes que hayan revisado los documentos.
        for lead in self:
            if type == 'tecnica':
                revisadas = lead.pt_revision_ids.filtered(lambda r: r.revisado)
                empleados = set(revisadas.mapped('employee_id.id'))
            else:
                revisadas = lead.pe_revision_ids.filtered(lambda r: r.revisado)
                empleados = set(revisadas.mapped('employee_ids.id'))
            
            if len(empleados) < 2:
                raise UserError(_('Debe haber al menos dos revisiones hechas por personas distintas para continuar.'))


    def action_advance_stage(self):
        if self.stage_name == 'Inscripción/Compra de bases':
            if self.bases_pay:
                if not self.oc_ids.filtered(lambda u: u.state != 'cancel' and u.type_purchase == 'bases'):
                    raise ValidationError('Favor de realizar la orden de compra antes de avanzar la etapa') 
                if not self.bases_doc:
                    raise ValidationError('Falta cargar las Bases correspondientes') 
            else:
                if not self.bases_doc:
                    raise ValidationError('Falta cargar las Bases correspondientes')
            if not self.bases_abstinencia_anticipo and self.bases_anticipo_porcentaje == 0:
                raise ValidationError('Falta agregar el % de anticipo')
            if self.bases_anticipo_porcentaje > 100:
                raise ValidationError('% de anticipo no valido')

        if self.stage_name == 'Visita de Obra':
            if not self.visita_acta:
                raise ValidationError('Falta cargar el Acta de la Visita')

        if self.stage_name == 'Junta de Aclaración de Dudas':
            if self.junta_line_ids:
                for junta in self.junta_line_ids:
                    if junta.asistente_ids and not junta.acta_junta:
                        raise ValidationError('Falta cargar el Acta de la Junta "%s" (Junta de Aclaración de Dudas).'% (junta.descripcion or junta.fecha_hora))

        if self.stage_name == 'Cotización de insumos y trabajos especiales':
            if not self.no_cotizar_insumos:
                if not self.input_ids:
                    raise ValidationError('Falta cargar información de los conceptos de trabajo y/o insumos. '
                                        'Si no es requerido cotizar, marque la casilla "No se necesita cotizar insumos".')

                count = len(self.oc_ids.filtered(lambda u: u.type_purchase == 'ins'))
                if self.input_ids and count == 0:
                    raise ValidationError('No se han creado cotizaciones de insumos. '
                                        'Si no es requerido cotizar, marque la casilla "No se necesita cotizar insumos".')

        if self.stage_name == 'Propuesta Técnica':
            self.check_doctos('tecnica')
            self.check_revisions('tecnica')

        if self.stage_name == 'Propuesta Económica':
            self.check_doctos('economica')
            if not self.pe_revision_ids.filtered(lambda r: r.autorizado):
                raise ValidationError('Debe tener al menos una revisión AUTORIZADA en Propuesta Económica para avanzar de etapa.')
            if not self.pe_presupuesto_autorizado:
                raise ValidationError('Debe AUTORIZAR el presupuesto de la propuesta económica para avanzar de etapa. '
                                    'Cargue los datos del documento Económico 2, marque como Revisado y luego Autorice.')

        if self.stage_name == 'Junta de Apertura de Propuestas':
            if not self.apertura_acta:
                raise ValidationError('Falta cargar el Acta de la Junta de Apertura de Propuestas.')

        if self.stage_name == 'Junta de Fallo':
            if not self.fallo_acta:
                raise ValidationError('Falta cargar el Acta de Fallo.')
            if not self.fallo_ganado:
                raise ValidationError(
                    'Debe marcar la casilla "Ganado" para avanzar a la siguiente etapa. '
                    'Si la licitación no fue ganada, use el botón "Marcar como Perdido".'
                )

            if not self.fallo_notif_directores_manual_sent and not self.fallo_notif_directores_auto_sent:
                raise ValidationError(
                    'Debe enviar la notificación a los directores antes de avanzar a etapa Ganado.'
                )

        sequence = self.stage_id.sequence
        reason = self.env['crm.revert.reason'].search([('name','=','Avance')])
        new_stage = self.env['crm.stage'].search([('sequence','=', sequence + 1)])

        if not new_stage:
            raise ValidationError('Existe un error en el flujo de las etapas, favor de revisar las configuraciones')

        self._log_stage_change(self.stage_id, new_stage, reason.id, 'Avance de etapa')
        self.stage_previous = self.stage_id.name
        old_stage_name = self.stage_name
        self.stage_id = new_stage.id
        if old_stage_name == 'Junta de Fallo' and self.fallo_ganado:
            return self.action_set_won_rainbowman()


    def action_revert_stage(self):
        # Abre el asistente para regresar la etapa con motivo.
        self.ensure_one()
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'crm.revert.stage.wizard',
            'view_mode': 'form',
            'target': 'new',
            'context': {'default_lead_id': self.id},}

    def action_generar_orden(self):
        oc_vals = self.get_orden_default_values(self.id)
        oc_vals_2 = oc_vals[:]
        oc_new = self._create_oc_async(oc_vals=oc_vals_2)
        return oc_new

    def get_orden_default_values(self, lead=False):
        if not self.partner_id.id:
            raise UserError('No se ha capturado información de contacto')

        if not self.origen_id.product_id.id:
            raise UserError('Falta configurar el concepto a pagar')

        if not self.env.user.has_group('project_extra.group_conv_authorizer'):
            raise UserError('Solo usuarios autorizadores pueden aprobar la compra de bases.')

        if not self.bases_pay:
            raise UserError('La licitación no necesita orden de compra')

        if not self.bases_cost or self.bases_cost == 0.00:
            raise UserError('Capturar el costo de la requisición.')

        if self.oc_ids.filtered(lambda u: u.state != 'cancel' and u.type_purchase == 'bases'):
            raise UserError('Ya existe Orden de pago ligada.')

        self.ensure_one()
        sequence = self.env['ir.sequence'].next_by_code('purchase.order')
        price = self.bases_cost
        
        orders = []
        order_lines = []
        taxes = []
        fpos = self.env['account.fiscal.position'].with_company(self.company_id)._get_fiscal_position(self.partner_id)
        taxes = self.origen_id.product_id.supplier_taxes_id._filter_taxes_by_company(self.company_id)

        name = self.origen_id.product_id.name
        if self.no_licitacion:
            name += ' ' + self.no_licitacion
        origin = 'Base - ' + self.name

        order_lines.append((0, 0, {
            'name': name, 'product_uom': self.origen_id.product_id.uom_po_id.id, 'product_id': self.origen_id.product_id.id, 'company_id': self.company_id.id,
            'partner_id': self.partner_id.id, 'currency_id': self.origen_id.product_id.currency_id.id, 'state': 'purchase', 'product_qty': 1, 
            'price_unit': price, 'taxes_id': fpos.map_tax(taxes) },))        
        values = {'lead_id': self.id, 'name': sequence, 'partner_id': self.partner_id.id, 'company_id': self.company_id.id, 'origin': origin,
            'currency_id': self.origen_id.product_id.currency_id.id, 'user_id': self.env.uid, 'state': 'purchase', 'invoice_status': 'no', 
            'type_purchase': 'bases', 'order_line': order_lines}
        orders.append(values)
        return orders


    def _create_oc_async(self, oc_vals):
        oc_obj = self.env['purchase.order']
        new_oc = oc_obj.create(oc_vals)        
        return new_oc

    def action_open_revert_logs(self):
        self.ensure_one()
        action = self.env.ref('project_extra.action_crm_revert_log').read()[0]
        action['domain'] = [('lead_id', '=', self.id)]
        action['context'] = {'default_lead_id': self.id}
        return action

    def action_send_bases(self):
        # Envía correo a group_conv_authorizer para solicitar autorización de compra de bases.
        for lead in self:
            if not lead.bases_cost:
                raise UserError('Debes capturar el costo de las bases.')

            correos_list = lead._get_authorizer_emails_from_group('project_extra.group_conv_authorizer')
            template = self.env.ref('project_extra.mail_tmpl_bases_solicitud', raise_if_not_found=False)

            if not correos_list:
                raise UserError(_('''No hay usuarios configurados con permiso para autorizar (o no tienen correo).
                        Agregue usuarios al grupo “Puede autorizar convocatorias”.'''))
            if not template:
                raise UserError(_('No se encontró la plantilla de correo para solicitud de bases.'))

            try:
                correos = ', '.join(correos_list)
                email_values = {'model': 'crm.lead', 'email_to': correos}
                template.send_mail(lead.id, force_send=True, email_values=email_values)
                lead.bases_notification_sent = True
                lead._post_html(_('Se envió correo a: ') + correos)
            except Exception:
                lead._post_html(_('Error al enviar el correo'))


    def action_authorize_bases(self):
        # La ejecuta un usuario del grupo project_extra.group_conv_authorizer. Genera OC y notifica a Finanzas.
        self.ensure_one()
        oc = self.action_generar_orden()
        correos_list = self._get_authorizer_emails_from_group('purchase.group_purchase_user')
        template = self.env.ref('project_extra.mail_tmpl_bases_autorizar', raise_if_not_found=False)

        if not correos_list:
            raise UserError('No hay usuarios configurados con permiso para autorizar (o no tienen correo).')
        if not template:
            raise UserError(_('No se encontró la plantilla de correo para solicitud de bases.'))

        try:
            correos = ', '.join(correos_list)
            email_values = {'model': 'purchase.order', 'email_to': correos}
            template.send_mail(oc.id, force_send=True, email_values=email_values)
            self.bases_notification_sent = True
            self._post_html(_('Se envió correo a: ') + correos)
        except Exception:
            self._post_html(_('Error al enviar el correo'))

        if self.stage_name == 'Declinado' and self.stage_previous == 'Inscripción/Compra de bases':
            dest_stage = self._get_stage_by_name('Inscripción/Compra de bases')
            self.write({'stage_id': dest_stage.id, 'stage_previous': self.stage_name})
            self._log_stage_change(self.stage_id, dest_stage, False, 'Autorización de compra de bases')


    def action_decline_bases(self):
        self.ensure_one()
        if not self.env.user.has_group('project_extra.group_conv_authorizer'):
            raise UserError(_('No tiene permisos para declinar la compra de bases.'))

        self.message_post(body=_('La compra de bases ha sido <b>DECLINADA</b> por %s.') % self.env.user.display_name)
        self.bases_pay = False
        self.bases_notification_sent = False
        self.bases_doc = False
        self.bases_doc_name = False

        dest_stage = self._get_stage_by_name('Declinado')
        if not dest_stage:
            raise UserError('No se encontró la etapa DECLINADO')

        old_stage = self.stage_id
        if old_stage.id != dest_stage.id:
            self.write({'stage_id': dest_stage.id, 'stage_previous': old_stage.name})

        self._log_stage_change(old_stage, dest_stage, False, 'Declinado')
        self._post_html(_('Declinada por %s.') % self.env.user.display_name, old_stage, dest_stage)


    def _get_emails(self):
        emails = set()
        for rec in self.stage_id.email_ids:
            if rec.work_email:
                emails.add(rec.work_email.strip())
            elif rec.private_email:
                emails.add(rec.private_email.strip())
            elif rec.address_id.email:
                emails.add(rec.address_id.email.strip())
            else:
                raise UserError(('No existen correos configurados del empleado %s') % rec.name)

        if self.stage_name == 'Visita de Obra':
            correos = self.visita_personas_ids
        elif self.stage_name == 'Junta de Fallo':
            correos = self.fallo_personas_ids
        else:
            correos = self.apertura_personas_ids

        if not correos:
            raise UserError(_('Debe asignar al menos una persona para la %s') % self.stage_name)

        for rec in correos:
            if rec.work_email:
                emails.add(rec.work_email.strip())
            elif rec.private_email:
                emails.add(rec.private_email.strip())
            elif rec.address_id.email:
                emails.add(rec.address_id.email.strip())
            else:
                raise UserError(('No existen correos configurados del empleado %s') % rec.name)

        return sorted(e for e in emails if '@' in e) 


    def _send_visita_reminder(self, manual=False):
        # Envía correo de recordatorio de visita usando la plantilla de visita.
        template = self.env.ref('project_extra.mail_tmpl_visita_recordatorio', raise_if_not_found=False)
        if not template:
            raise UserError('No se encontró la plantilla de correo para recordatorio de visita de obra.')

        correos_list = self._get_emails()
        correos = ', '.join(correos_list)

        for lead in self:
            if not lead.visita_fecha:
                raise UserError(_('Debe capturar la fecha de la visita de obra.'))
            if not lead.visita_personas_ids:
                raise UserError(_('Debe asignar al menos una persona para la Visita de Obra.'))

            email_values = {'email_to': correos,}
            template.send_mail(lead.id, force_send=True, email_values=email_values)
            if manual:
                lead.visita_notif_manual_sent = True
            else:
                lead.visita_notif_auto_sent = True

            lead.message_post(body=_("Se envió recordatorio de visita de obra a: %s") % correos)

    def action_send_visita_reminder(self):
        #Botón manual para enviar recordatorio de visita.
        self.ensure_one()
        if not self.visita_fecha:
            raise UserError('Debe capturar la fecha de visita.')
        if not self.visita_personas_ids:
            raise UserError(_('Debe asignar al menos una persona para la Visita de Obra.'))
        self._send_visita_reminder(manual=True)

    def _send_junta_reminder(self, manual=False):
        template = self.env.ref('project_extra.mail_tmpl_junta_recordatorio', raise_if_not_found=False)
        if not template:
            raise UserError(_('No se encontró la plantilla de correo para recordatorio de la Junta de Aclaración de Dudas.'))

        for lead in self:
            email_to = ', '.join(lead._get_emails())
            email_values = {'email_to': email_to}

            template.send_mail(lead.id, force_send=True, email_values=email_values)

            if manual:
                lead.junta_notif_manual_sent = True
            else:
                lead.junta_notif_auto_sent = True

            lead.message_post(body=_("Se envió recordatorio de junta a: %s") % email_to)
            

    def action_send_junta_reminder_manual(self):
        self.ensure_one()
        # Si hay líneas de junta, usar el nuevo sistema
        if self.junta_line_ids:
            today = fields.Date.context_today(self)
            
            # Filtrar juntas con fecha >= hoy que no hayan sido notificadas manualmente
            juntas_pendientes = self.junta_line_ids.filtered(lambda j: j.fecha_hora and j.fecha_hora.date() >= today and not j.notif_manual 
                and j.asistente_ids)
            
            if not juntas_pendientes:
                raise UserError(_('No hay juntas pendientes de notificar. '
                                'Las juntas deben tener fecha igual o posterior a hoy, asistentes asignados y no haber sido notificadas previamente.'))
            
            errores = []
            enviados = 0
            for junta in juntas_pendientes:
                try:
                    junta._send_junta_line_reminder(manual=True)
                    enviados += 1
                except Exception as e:
                    errores.append(f"{junta.descripcion or 'Junta sin descripción'}: {str(e)}")
            
            if errores:
                self.message_post(body=_('Se enviaron %d notificaciones. Errores: %s') % (enviados, ', '.join(errores)))
                raise UserError(_('Se enviaron %d notificaciones con algunos errores:\n%s') % (enviados, '\n'.join(errores)))
            
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': _('Notificaciones enviadas'),
                    'message': _('Se enviaron %d notificaciones correctamente.') % enviados,
                    'type': 'success',
                    'sticky': False,}}
        
        self._send_junta_reminder(manual=True)


    def _send_junta_dudas_deadline_reminder(self):
        template = self.env.ref('project_extra.mail_tmpl_junta_dudas_deadline', raise_if_not_found=False)
        if not template:
            raise UserError(_('No se encontró la plantilla de recordatorio de fecha límite de dudas.'))

        for lead in self:
            if not lead.junta_obligaboria: # Pendiente tarea 0050    (lead.junta_obligatoria and lead.junta_fecha_limite_dudas):
                continue

            if lead.stage_name == 'Junta de Aclaración de Dudas':
                correos = ', '.join(lead._get_emails())
                template.send_mail(lead.id, force_send=True, email_values={'email_to': correos})
                lead.junta_dudas_notif_auto_sent = True
                lead.message_post(body=_("Se envió recordatorio de FECHA LÍMITE DE DUDAS a: %s") % correos)


    def _send_apertura_reminder(self, manual=False):
        template = self.env.ref('project_extra.mail_tmpl_apertura_recordatorio', raise_if_not_found=False)
        if not template:
            raise UserError(_('No se encontró la plantilla de correo para recordatorio de junta de apertura de propuestas.'))

        for lead in self:
            if not lead.apertura_fecha:
                raise UserError(_('Debe capturar la fecha de la junta de apertura de propuestas.'))
            if not lead.apertura_personas_ids:
                raise UserError(_('Debe asignar al menos una persona para la Junta de Apertura de Propuestas.'))

            correos = ', '.join(lead._get_emails())
            template.send_mail(lead.id, force_send=True, email_values={'email_to': correos})
            lead.write({'apertura_notif_manual_sent': manual, 'apertura_notif_auto_sent': not manual})
            lead.message_post(body=_("Se envió recordatorio de junta de apertura de propuestas a: %s") % correos)


    def action_send_apertura_reminder(self):
        # Botón manual para enviar recordatorio de junta de apertura de propuestas.
        self.ensure_one()
        if not self.apertura_fecha:
            raise UserError(_('Debe capturar la fecha de la junta de apertura de propuestas.'))
        if not self.apertura_personas_ids:
            raise UserError(_('Debe asignar al menos una persona a la junta de apertura de propuestas.'))

        self._send_apertura_reminder(manual=True)


    def _send_fallo_notification(self, manual=False):
        template = self.env.ref('project_extra.mail_tmpl_fallo_ganado', raise_if_not_found=False)
        if not template:
            raise UserError(_('No se encontró la plantilla de correo para fallo GANADO.'))

        for lead in self:
            # Si quieres forzar que solo se mande en etapa JUNTA DE FALLO:
            if lead.stage_name != 'Junta de Fallo':
                continue

            correos = ', '.join(lead._get_emails())
            template.send_mail(lead.id, force_send=True, email_values={'email_to': correos})
            if manual:
                lead.fallo_notif_manual_sent = True
            else:
                lead.fallo_notif_auto_sent = True

            lead.message_post(body=_("Se envió notificación de FALLO GANADO."))

            
    def action_send_fallo_notification(self):
        self.ensure_one()
        self._send_fallo_notification(manual=True)

    def action_send_fallo_directores_notification(self):
        # Botón manual para enviar notificación de FALLO GANADO a directores.
        self.ensure_one()
        if not self.fallo_ganado:
            raise UserError(_('Debe marcar la casilla "Ganado" antes de enviar la notificación a directores.'))

        # Obtener correos de directores
        correos_list = self._get_authorizer_emails_from_group('project_extra.group_conv_authorizer')
        template = self.env.ref('project_extra.fallo_ganado_mail_tmpl_directores', raise_if_not_found=False)

        if not template:
            self._post_html(_('No se encontró la plantilla de correo para notificar a directores.'))
            raise UserError(_('No se encontró la plantilla de correo para notificar a directores.'))

        if not correos_list:
            self._post_html(_('No hay directores configurados con correo electrónico.'))
            raise UserError(_('No hay directores configurados con correo electrónico.'))

        try:
            correos = ', '.join(correos_list)
            email_values = {'model': 'crm.lead', 'email_to': correos}
            template.send_mail(self.id, force_send=True, email_values=email_values)
            self.write({'fallo_notif_directores_manual_sent': True})
            self._post_html(_('Notificación de fallo ganado enviada a directores: ') + correos)
        except Exception as e:
            _logger.error("Error al enviar notificación de fallo a directores: %s", str(e))
            self._post_html(_('Error al enviar el correo a directores'))
            raise UserError(_('Error al enviar el correo a directores: %s') % str(e))

        
    @api.model
    def cron_send_visita_reminders(self):
        # Cron: envía recordatorio un día antes de la fecha de visita.
        today = fields.Date.context_today(self)
        target = today + relativedelta(days=1)

        domain = [('visita_obligatoria','=',True), ('visita_fecha','=',target), ('visita_notif_auto_sent','=',False), ('stage_name','=','Visita de Obra')]
        leads = self.search(domain)
        for rec in leads:
            rec._send_visita_reminder(manual=False)


    @api.model
    def cron_send_junta_reminders(self):
        # Cron: envía recordatorio un día antes de la Junta de Aclaración de Dudas.
        today = fields.Date.context_today(self)
        target = today + relativedelta(days=1)

        domain = [('junta_obligatoria','=',True), ('junta_notif_auto_sent','=',False), #Pendiente 0050 ('junta_fecha','=',target),
            ('stage_name','=','Junta de Aclaración de Dudas')]
        leads = self.search(domain)
        for rec in leads:
            rec._send_junta_reminder(manual=False)


    @api.model
    def cron_send_junta_dudas_deadline(self):
        # Cron: envía recordatorio un día antes de la fecha límite de dudas.
        today = fields.Date.context_today(self)
        target = today + relativedelta(days=1)
        domain = [('junta_obligatoria','=',True), ('junta_dudas_notif_auto_sent','=',False), #Pendietne 0050 ('junta_fecha_limite_dudas','=',target),
            ('stage_name','=','Junta de Aclaración de Dudas')]
        leads = self.search(domain)
        for rec in leads:
            rec._send_junta_dudas_deadline_reminder()

    @api.model
    def cron_send_junta_line_reminders(self):
        """
        T0050 - Cron: envía recordatorio de Juntas de Aclaración de Dudas.
        Solo envía a juntas con fecha IGUAL O POSTERIOR a hoy que no hayan sido notificadas.
        NO envía a juntas con fechas pasadas.
        """
        today = fields.Date.context_today(self)
        
        _logger.info("Ejecutando cron_send_junta_line_reminders para fecha >= %s", today)
        
        leads = self.search([
            ('junta_obligatoria', '=', True),
            ('stage_name', '=', 'Junta de Aclaración de Dudas')
        ])
        
        enviados = 0
        errores = 0
        
        for lead in leads:
            for junta_line in lead.junta_line_ids:
                # Solo juntas con fecha >= hoy, no notificadas automáticamente y con asistentes
                if (junta_line.fecha_hora and 
                    junta_line.fecha_hora.date() >= today and 
                    not junta_line.notif_auto and
                    junta_line.asistente_ids):
                    try:
                        junta_line._send_junta_line_reminder(manual=False)
                        enviados += 1
                    except Exception as e:
                        errores += 1
                        _logger.error(
                            "Error al enviar recordatorio automático de junta (Lead: %s, Junta: %s): %s",
                            lead.id, junta_line.id, str(e)
                        )
        
        _logger.info("Cron junta line finalizado: %d enviados, %d errores", enviados, errores)

    @api.model
    def cron_send_junta_line_deadline_reminders(self):
        """
        T0050 - Cron: envía recordatorio de fecha límite de preguntas.
        Solo envía a juntas con fecha límite IGUAL O POSTERIOR a hoy que no hayan sido notificadas.
        NO envía a juntas con fechas pasadas.
        """
        today = fields.Date.context_today(self)
        
        _logger.info("Ejecutando cron_send_junta_line_deadline_reminders para fecha >= %s", today)
        
        template = self.env.ref('project_extra.mail_tmpl_junta_line_deadline', raise_if_not_found=False)
        if not template:
            _logger.warning("No se encontró la plantilla mail_tmpl_junta_line_deadline")
            return
        
        leads = self.search([
            ('junta_obligatoria', '=', True),
            ('stage_name', '=', 'Junta de Aclaración de Dudas')
        ])
        
        enviados = 0
        
        for lead in leads:
            for junta_line in lead.junta_line_ids:
                # Solo juntas con fecha límite >= hoy y con asistentes
                if (junta_line.fecha_limite_preguntas and 
                    junta_line.fecha_limite_preguntas.date() >= today and
                    junta_line.asistente_ids):
                    try:
                        emails = junta_line._get_asistente_emails()
                        if emails:
                            email_to = ', '.join(emails)
                            template.send_mail(junta_line.id, force_send=True, email_values={'email_to': email_to})
                            lead.message_post(
                                body=_("Se envió recordatorio de FECHA LÍMITE DE PREGUNTAS para junta '%s' a: %s") 
                                    % (junta_line.descripcion or 'Sin descripción', email_to)
                            )
                            enviados += 1
                    except Exception as e:
                        _logger.error(
                            "Error al enviar recordatorio de fecha límite (Lead: %s, Junta: %s): %s",
                            lead.id, junta_line.id, str(e)
                        )
        
        _logger.info("Cron fecha límite finalizado: %d enviados", enviados)
    
    def _sync_pt_doc_lines(self):
        # Sincroniza la tabla de Propuesta Técnica con los documentos seleccionados en ptdcto_ids.
        docline = self.env['crm.propuesta.tecnica.doc']
        for lead in self:
            docs = lead.ptdcto_ids
            existing_by_doc = {line.docto_id.id: line for line in lead.pt_doc_line_ids}
            wanted_ids = set(docs.ids)

            # Crear líneas nuevas para documentos que faltan
            for doc in docs:
                if doc.id not in existing_by_doc:
                    docline.create({'lead_id': lead.id, 'docto_id': doc.id,})

            # Eliminar líneas de documentos que ya no están seleccionados
            for line in lead.pt_doc_line_ids:
                if line.docto_id.id not in wanted_ids:
                    line.unlink()


    @api.onchange('ptdcto_ids')
    def _onchange_ptdcto_ids(self):
        if self.id:
            self._sync_pt_doc_lines()

    def action_cargar_insumos(self):
        for record in self:
            if not record.input_file:
                raise ValidationError('Seleccione un archivo para cargar.')

            if record.input_file and record.input_ids:
                raise ValidationError('Ya hay información cargada. En caso de ser necesario volver a cargar debe eliminarlos.')

            filename, file_extension = os.path.splitext(record.input_filename)
            if file_extension in ['.xlsx', '.xls', '.xlsm']:
                record.__leer_carga_insumos()
            else:
                raise ValidationError('Seleccione un archivo tipo xlsx, xls, xlsm')


    def __leer_carga_insumos(self):
        for record in self:
            file = tempfile.NamedTemporaryFile(suffix=".xlsx")
            file.write(binascii.a2b_base64(record.input_file))
            file.seek(0)
            xlsx_file = file.name
            
            wb = openpyxl.load_workbook(filename=xlsx_file, data_only=True)
            sheets = wb.sheetnames
            sheet_name = sheets[0]
            sheet = wb[sheet_name]
            registros = []
            cargar = False
            for row in sheet.iter_rows(values_only=True):
                col1 = str(row[0]).strip()
                col2 = str(row[1]).strip()
                col3 = str(row[2]).strip()
                col4 = str(row[3]).strip()
                col5 = str(row[4]).strip()
                col6 = str(row[5]).strip()
                col7 = str(row[6]).strip()
                col8 = str(row[7]).strip()
                if col1 == 'None':
                    col1 = ''
                if col2 == 'None':
                    col2 = ''
                if col3 == 'None':
                    col3 = ''
                if col4 == 'None':
                    col4 = ''
                if col5 == 'None':
                    col5 = ''
                if col6 == 'None':
                    col6 = ''
                if col7 == 'None':
                    col7 = ''
                if col8 == 'None':
                    col8 = ''
                if col1.upper() == 'CÓDIGO':
                    cargar = True

                if cargar:
                    if col1 != '' and col8 != '' and col1.upper() != 'CÓDIGO':
                        registro = {'col1': col1, 'col2': col2, 'col3': col3, 'col4': col4, 'col5': col5, 'col6': col6, 'col7': col7, 'col8': col8}
                        registros.append((0, 0, registro))

            record.write({'input_ids': registros})


    def action_genera_insumos(self):
        self.env.cr.execute('SELECT col1, COUNT(*) num FROM crm_input_line WHERE lead_id = ' + str(self.id) + ' GROUP BY 1 HAVING COUNT(*) > 1')
        duplicado = self.env.cr.dictfetchall()
        if duplicado:
            raise UserError('Existen conceptos repetidos favor de revisar el archivo.')

        self.env.cr.execute('''SELECT ID min_id, (case when UPPER(col3) = 'UNIDAD' then 'col3' else 'col5' end) unidad, 
                (case when UPPER(col5) = 'CANTIDAD' then 'col5' else 'col6' end) cantidad, 
                (case when UPPER(col6) in ('PRECIO', 'COSTO UNITARIO') then 'col6' else 'col7' end) precio 
            FROM crm_input_line ci WHERE ci.id = (select MIN(ID) min_id from crm_input_line ci WHERE ci.lead_id = ''' + str(self.id) + ')')
        min_id = self.env.cr.dictfetchall()

        unidad = min_id[0]['unidad']
        cantidad = min_id[0]['cantidad']
        precio = min_id[0]['precio']

        self.env.cr.execute('''UPDATE crm_input_line cil SET input_ex = True, input_id = t1.IDCOD
            FROM (SELECT cil.id, TRIM(cil.col1) code, MIN(pt.ID) idcod, COUNT(pt.id) num 
                    FROM crm_input_line cil LEFT JOIN product_template pt ON TRIM(cil.col1) = pt.default_code 
                   WHERE cil.LEAD_ID = ''' + str(self.id) + ' and cil.id != ' + str(min_id[0]['min_id']) + ''' AND cil.input_ex = False GROUP BY 1, 2) as t1
            WHERE cil.id = t1.id AND t1.num != 0;
            UPDATE crm_input_line cil SET account_ex = true
            FROM (SELECT cil.id, TRIM(cil.col1) code, COUNT(pt.property_account_expense_id) num 
                    FROM crm_input_line cil JOIN product_template pt ON TRIM(cil.col1) = pt.default_code 
                   WHERE cil.LEAD_ID = ''' + str(self.id) + ' AND cil.id != ' + str(min_id[0]['min_id']) + ''' AND pt.property_account_expense_id IS NOT NULL
                   GROUP BY 1, 2) as t1
            WHERE cil.id = t1.id;
            UPDATE crm_input_line cil SET tipinsumo_ex = true
            FROM (SELECT cil.id, TRIM(cil.col1) code, COUNT(pt.tipo_insumo_id) num 
                    FROM crm_input_line cil JOIN product_template pt ON TRIM(cil.col1) = pt.default_code 
                   WHERE cil.LEAD_ID = ''' + str(self.id) + ' AND cil.id != ' + str(min_id[0]['min_id']) + ''' AND pt.tipo_insumo_id IS NOT NULL
                   GROUP BY 1, 2) as t1
            WHERE cil.id = t1.id; ''')

        iva = self.env['account.tax'].search([('name','=','16%'),('type_tax_use','=','purchase')])
        for rec in self.input_ids.filtered(lambda u: not u.input_ex):
            if rec.id != min_id[0]['min_id']:
                statement = ('''SELECT cil.col1 code, cil.col2 name, uu.id uom, pc.id cat,
                        (CASE WHEN uc.NAME->>'en_US' = 'Service' THEN 'service' ELSE 'consu' END) type, ''' + cantidad + ' qty, ' + precio + 
                        '::float importe FROM crm_input_line cil JOIN uom_uom uu ON (CASE WHEN cil.' + unidad + 
                        " IN ('%MO', 'PIE TAB', '%') THEN 'pza' ELSE lower(cil." + unidad + ''') END) = lower(uu.name->>'en_US') 
                                JOIN uom_category uc ON uu.CATEGORY_ID = uc.ID 
                                JOIN product_category pc ON pc.NAME = 'All'
                    WHERE cil.id = ''' + str(rec.id))
                self.env.cr.execute(statement)
                info = self.env.cr.dictfetchall()
                if info:                
                    insumo = self.env['product.template'].create({'categ_id': info[0]['cat'], 'uom_id': info[0]['uom'], 'uom_po_id': info[0]['uom'], 
                        'type': info[0]['type'], 'default_code': info[0]['code'], 'name': info[0]['name'], 'purchase_ok': True, 'sale_ok': False, 
                        'supplier_taxes_id': [(6, 0, iva.ids)], 'standard_price': info[0]['importe'], 'active': True,})
                    rec.write({'input_ex': True, 'input_id': insumo.id})


    def action_genera_cotizaciones(self):
        self.ensure_one()
        count = len(self.input_ids.filtered(lambda u: not u.input_ex))
        if count != 1:
            raise UserError('Existen insumos sin cargar, revise la información.')

        count = len(self.input_ids.filtered(lambda u: not u.account_ex))
        if count != 1:
            raise UserError('Los insumos cargados no cuentan con la información contable. Favor de contactar con el área correspondiente')

        count = len(self.input_ids.filtered(lambda u: not u.tipinsumo_ex))
        if count != 1:
            raise UserError('Los insumos cargados no cuentan con el tipo. Favor de contactar con el área correspondiente')

        return {
            'type': 'ir.actions.act_window',
            'res_model': 'crm.cotizacion.insumos.wizard',
            'view_mode': 'form',
            'target': 'new',
            'context': {'default_lead_id': self.id}}


    def action_unlink_insumos(self):
        for record in self:
            record.input_ids.unlink()

    def action_cargar_concept(self):
        for record in self:
            if not record.doctoconcept_id:
                raise ValidationError('Falta agregar el tipo de archivo.')

            docto = self.env['documents.document'].search([('res_model','=','crm.lead'), ('res_id','=',record.id), 
                ('file_extension','in',['xlsx', 'xls', 'xlsm']), '|', ('name','ilike','E02'), ('name','ilike','Economico 2')])
            if not docto:
                raise ValidationError('El documento E02 no se encuentra cargado, favor de revisar la documentación.')

            if record.concept_ids:
                raise ValidationError('Ya hay información cargada. En caso de ser necesario volver a cargar debe eliminarlos.')

            if docto.file_extension in ['xlsx', 'xls', 'xlsm']:
                record.__leer_carga_concept(docto)
            else:
                raise ValidationError('Seleccione un archivo tipo xlsx, xls, xlsm')


    def __leer_carga_concept(self, docto):
        for record in self:
            if not record.doctoconcept_id:
                raise ValidationError('Falta agregar el tipo de archivo.')

            for x in record.doctoconcept_id:
                inicio = x.inicio_datos
                for y in x.configdoc_ids:
                    if y.no_columna != '*':
                        if y.columna == 'codigo':
                            cod = int(y.no_columna) - 1
                        if y.columna == 'descripcion':
                            desc = int(y.no_columna) - 1
                        if y.columna == 'unidad':
                            uni = int(y.no_columna) - 1
                        if y.columna == 'precio_unitario':
                            prec = int(y.no_columna) - 1
                        if y.columna == 'cantidad':
                            qty = int(y.no_columna) - 1
                        if y.columna == 'importe':
                            imp = int(y.no_columna) - 1

            attachment = docto.attachment_id
            decoded_data = base64.b64decode(attachment.datas)
            workbook = openpyxl.load_workbook(filename=io.BytesIO(decoded_data), data_only=True)
            sheet = workbook.active
            column = sheet.max_column
            registros = []
            partidas = []
            cargar = True
            partida = False
            contador = 1
            for row in sheet.iter_rows(values_only=True):
                if contador < inicio:
                    contador += 1
                    continue

                if row[desc] != None:
                    if row[desc].upper() == 'RESUMEN DE PARTIDAS':
                        cargar = False
                        partida = True

                if cargar:
                    if row[cod] == row[desc] == row[uni] == row[prec] == row[qty] == row[imp] == None:
                        _logger.warning('No se guarda la información')
                    else:
                        registro = {'col1': row[cod], 'col2': row[desc], 'col3': row[uni], 'col4': row[qty], 'col5': row[prec], 'col6': row[imp]}
                        registros.append((0, 0, registro))

                if partida:
                    if row[cod] != None and row[desc] != None:
                        registro = {'col1': row[cod], 'col2': row[desc]}
                        partidas.append((0, 0, registro))

            record.write({'concept_ids': registros, 'budget_ids': partidas})


    def action_genera_partidas(self):
        count = len(self.budget_ids.filtered(lambda u: not u.budget_id))
        if count == 0:
            raise UserError('Las partidas fueron cargadas correctamente, favor de continuar con la carga de conceptos')
        else:
            self.env.cr.execute('SELECT carga_partidas(' + str(self.id)+ ', ' + str(self.env.user.id) + ')')
            partidas = self.env.cr.dictfetchall()

    def action_genera_concept(self):
        count = len(self.budget_ids.filtered(lambda u: not u.budget_id))
        if count != 0:
            raise UserError('Las partidas no han sido cargadas, favor de realizar la carga')

        iva = self.env['account.tax'].search([('amount','=',16), ('type_tax_use','=','sale')])
        partida = 0
        for rec in self.concept_ids.filtered(lambda u: not u.concept_ex):
            if not rec.concept_ex:
                partida_id = self.env['crm.budget.line'].search([('lead_id','=',rec.lead_id.id), ('col1','=',rec.col1)])
                if partida_id:
                    partida = partida_id.budget_id.id

                if partida != 0 and rec.col1 != '' and rec.col1 != partida_id.budget_id.code:
                    concept_id = self.env['product.template'].search([('budget_id','=',partida), ('default_code','=',rec.col1)])
                    if concept_id:
                        if concept_id.property_account_income_id:
                            rec.write({'concept_ex': True, 'concept_id': concept_id.id, 'account_ex': True})
                        else:
                            rec.write({'concept_ex': True, 'concept_id': concept_id.id})
                    else:
                        statement = ('''SELECT cil.col1 code, cil.col2 name, uu.id uom, pc.id cat, REPLACE(cil.col4, ',', '') qty, 
                                (CASE WHEN cil.col5 = '' THEN '0.0' ELSE REPLACE(cil.col5, ',', '') END)::float importe 
                            FROM crm_concept_line cil JOIN uom_uom uu ON lower(cil.col3) = lower(uu.name->>'en_US') JOIN product_category pc ON pc.NAME = 'All' 
                            WHERE cil.id = ''' + str(rec.id))
                        self.env.cr.execute(statement)
                        info = self.env.cr.dictfetchall()
                        if info:
                            insumo = self.env['product.template'].create({'categ_id': info[0]['cat'], 'uom_id': info[0]['uom'], 'uom_po_id': info[0]['uom'], 
                                'type': 'service', 'default_code': info[0]['code'], 'name': info[0]['name'], 'purchase_ok': False, 'sale_ok': True, 
                                'taxes_id': [(6, 0, iva.ids)], 'standard_price': info[0]['importe'], 'list_price': info[0]['importe'], 
                                'service_tracking': 'task_in_project', 'active': True, 'budget_id': partida})
                            rec.write({'concept_ex': True, 'concept_id': insumo.id})

        for rec in self.concept_ids.filtered(lambda u: u.concept_ex):
            if rec.concept_id.property_account_income_id:
                rec.write({'account_ex': True})

    def action_cargar_presupuesto_economico(self):
        self.ensure_one()
        docto = self.env['documents.document'].search([('res_model','=','crm.lead'), ('res_id','=',self.id), 
                ('file_extension','in',['xlsx', 'xls', 'xlsm']), '|', ('name','ilike','E02'), ('name','ilike','Economico 2')])
        if not docto:
            raise ValidationError('El documento E02 no se encuentra cargado, favor de revisar la documentación.')

        attachment = docto.attachment_id
        if not attachment or not attachment.datas:
            raise ValidationError(_('El documento Económico 2 no tiene contenido.'))

        decoded_data = base64.b64decode(attachment.datas)
        workbook = openpyxl.load_workbook(filename=io.BytesIO(decoded_data), data_only=True)
        sheet = workbook.active
        subtotal, iva, total = 0.0, 0.0, 0.0

        for row in range(len([row for row in sheet if any(cell.value is not None for cell in row)]), sheet.max_row + 1):
            for col in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row, column=col)
                if cell.value and isinstance(cell.value, str):
                    val_upper = cell.value.upper().strip()
                    value_cell = sheet.cell(row=row, column=col + 1)
                    
                    if 'SUBTOTAL' in val_upper and 'M. N.' in val_upper:
                        if value_cell.value and isinstance(value_cell.value, (int, float)):
                            subtotal = float(value_cell.value)
                    elif 'IVA' in val_upper and 'M. N.' in val_upper:
                        if value_cell.value and isinstance(value_cell.value, (int, float)):
                            iva = float(value_cell.value)
                    elif 'TOTAL' in val_upper and 'M. N.' in val_upper and 'SUBTOTAL' not in val_upper:
                        if value_cell.value and isinstance(value_cell.value, (int, float)):
                            total = float(value_cell.value)
        
        if subtotal ==  iva ==  total == 0.0:
            raise ValidationError(_('No se encontraron los valores de SUBTOTAL, IVA y TOTAL. Verifique que el archivo tenga el formato correcto.'))
        
        self.write({'pe_presupuesto_subtotal': subtotal, 'pe_presupuesto_iva': iva, 'pe_presupuesto_total': total,})
        

    def action_autorizar_presupuesto(self):
        self.ensure_one()
        if not self.pe_presupuesto_revisado:
            raise ValidationError(_('Debe marcar el campo "Revisado" antes de poder autorizar el presupuesto.'))
        if self.pe_presupuesto_total <= 0:
            raise ValidationError(_('Debe cargar los datos del presupuesto antes de autorizar. '
                            'Use el botón "Cargar datos" para extraer la información del documento Económico 2.'))
        
        self.write({'pe_presupuesto_autorizado': True})
        """return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Presupuesto Autorizado'),
                'message': _('El presupuesto ha sido autorizado correctamente.'),
                'type': 'success',
                'sticky': False,}} """


    def _check_pe_authorization_complete(self):
        self.ensure_one()
        if not self.pe_presupuesto_autorizado:
            return False
        
        revisiones_autorizadas = self.pe_revision_ids.filtered(lambda r: r.autorizado)
        if not revisiones_autorizadas:
            return False
        return True


    def action_cargar_basicos(self):
        for record in self:
            if not record.doctobasicos_id:
                raise ValidationError('Falta agregar el tipo de archivo.')

            if not record.combo_ids:
                raise ValidationError('Cargar primero la información del archivo E09.')

            docto = self.env['documents.document'].search([('res_model','=','crm.lead'), ('res_id','=',record.id), 
                ('file_extension','in',['xlsx', 'xls', 'xlsm']), '|', ('name','ilike','E10 B'), ('name','ilike','Economico 10 A')])
            if not docto:
                raise ValidationError('El documento E10 no se encuentra cargado, favor de revisar la documentación.')

            if record.basico_ids:
                raise ValidationError('Ya hay información cargada. En caso de ser necesario volver a cargar debe eliminarlos.')

            if docto.file_extension in ['xlsx', 'xls', 'xlsm']:
                record.__leer_carga_basico(docto)
            else:
                raise ValidationError('Seleccione un archivo tipo xlsx, xls, xlsm')


    def __leer_carga_basico(self, docto):
        for record in self:
            if not record.doctobasicos_id:
                raise ValidationError('Se debe de seleccionar el tipo de documento a cargar')

            for x in record.doctobasicos_id:
                inicio = x.inicio_datos
                for y in x.configdoc_ids:
                    if y.no_columna != '*':
                        if y.columna == 'concepto':
                            conc = int(y.no_columna) - 1
                        if y.columna == 'codigo':
                            cod = int(y.no_columna) - 1
                        if y.columna == 'descripcion':
                            desc = int(y.no_columna) - 1
                        if y.columna == 'unidad':
                            uni = int(y.no_columna) - 1
                        if y.columna == 'precio_unitario':
                            prec = int(y.no_columna) - 1
                        if y.columna == 'cantidad':
                            qty = int(y.no_columna) - 1
                        if y.columna == 'importe':
                            imp = int(y.no_columna) - 1
                if x.docto_rel:
                    iniciorel = x.docto_rel.inicio_datos
                    for y in x.docto_rel.configdoc_ids:
                        if y.no_columna != '*':
                            if y.columna == 'codigo':
                                codrel = int(y.no_columna) - 1
                            if y.columna == 'descripcion':
                                descrel = int(y.no_columna) - 1
                            if y.columna == 'unidad':
                                unirel = int(y.no_columna) - 1
                            if y.columna == 'precio_unitario':
                                precrel = int(y.no_columna) - 1

            doctos = self.env['documents.document'].search([('res_model','=','crm.lead'), ('res_id','=',record.id), 
                ('file_extension','in',['xlsx', 'xls', 'xlsm']), ('name','ilike','10 Relac')])
            attachment = doctos.attachment_id
            decoded_data = base64.b64decode(attachment.datas)
            workbook = openpyxl.load_workbook(filename=io.BytesIO(decoded_data), data_only=True)
            sheet = workbook.active
            registros = []
            contador = 1
            for row in sheet.iter_rows(values_only=True):
                if contador < inicio:
                    contador += 1
                    continue

                if row[codrel] == None:
                    continue

                registro = {'col1': row[cod], 'col2': row[desc], 'col3': row[uni], 'col4': row[prec]}
                registros.append((0, 0, registro))
            record.write({'relacion_ids': registros})

            attachment = docto.attachment_id
            decoded_data = base64.b64decode(attachment.datas)
            workbook = openpyxl.load_workbook(filename=io.BytesIO(decoded_data), data_only=True)
            sheet = workbook.active
            registros = []
            contador = 1
            for row in sheet.iter_rows(values_only=True):
                if contador < inicio:
                    contador += 1
                    continue

                if row[conc] == None or row[conc].replace(' ', '') == '':
                    continue

                relacion = self.env['crm.basico.relacion'].search([('lead_id','=',record.id), ('col1','=',row[conc])])
                if relacion:
                    relacion_id = relacion
                    basico = row[conc]
                    continue

                if row[qty] == None:
                    continue
                if len(str(row[qty]).replace(' ', '')) == 0:
                    continue
                if not isinstance(row[imp], (int, float)):
                    continue
                
                registro = {'relacion_id': relacion_id.id, 'col1': row[cod], 'col2': row[desc], 'col3': row[uni], 'col4': row[prec], 'col5': row[qty], 
                    'col6': row[imp], 'basico': basico}
                registros.append((0, 0, registro))
            record.write({'basico_ids': registros})


    def action_generar_basicos(self):
        count = len(self.basico_ids.filtered(lambda u: not u.combo_ex))
        if count == 0:
            raise ValidationError('Los basicos fueron generados correctamente, favor de continuar con la generación de las tarjetas de conceptos')
        else:
            self.env.cr.execute('SELECT cbr.COL1, cbr.COL2 FROM crm_basico_line cbr WHERE cbr.LEAD_ID = ' + str(self.id) + ''' AND cbr.COMBO_EX IS FALSE 
                AND NOT EXISTS(SELECT * FROM crm_basico_relacion cbr2 WHERE cbr.COL1 = cbr2.COL1 AND cbr.LEAD_ID = cbr2.LEAD_ID) 
                AND NOT EXISTS(SELECT * FROM product_template pt WHERE pt.DEFAULT_CODE = cbr.COL1) GROUP BY 1, 2 ORDER BY 1''')
            ex_basicos = self.env.cr.dictfetchall()
            mensaje = ''
            for x in ex_basicos:
                mensaje += x['col1'] + ' - ' + x['col2'] + '\n'
            if mensaje != '':
                raise ValidationError('Faltan de cargar los siguientes insumos de basicos: %s' % mensaje)
            self.env.cr.execute('SELECT generar_basicos(' + str(self.id) + ', ' + str(self.env.user.id) + ')')
            basicos = self.env.cr.dictfetchall()

    def action_cargar_combo(self):
        for record in self:
            if not record.doctomatriz_id:
                raise ValidationError('Falta agregar el tipo de archivo.')

            docto = self.env['documents.document'].search([('res_model','=','crm.lead'), ('res_id','=',record.id), 
                ('file_extension','in',['xlsx', 'xls', 'xlsm']), '|', ('name','ilike','E09'), ('name','ilike','Economico 9')])
            if not docto:
                raise ValidationError('El documento E09 no se encuentra cargado, favor de revisar la documentación.')

            if record.combo_ids:
                raise ValidationError('Ya hay información cargada. En caso de ser necesario volver a cargar debe eliminarlos.')

            if docto.file_extension in ['xlsx', 'xls', 'xlsm']:
                record.__leer_carga_combo(docto)
            else:
                raise ValidationError('Seleccione un archivo tipo xlsx, xls, xlsm')


    def __leer_carga_combo(self, docto):
        for record in self:
            if not record.doctomatriz_id:
                raise ValidationError('Se debe de seleccionar el tipo de documento a cargar')

            inicio = record.doctomatriz_id.inicio_datos
            for x in record.doctomatriz_id.configdoc_ids:
                if x.no_columna != '*':
                    if x.columna == 'concepto':
                        conc = int(x.no_columna) - 1
                    if x.columna == 'codigo':
                        cod = int(x.no_columna) - 1
                    if x.columna == 'descripcion':
                        desc = int(x.no_columna) - 1
                    if x.columna == 'unidad':
                        uni = int(x.no_columna) - 1
                    if x.columna == 'precio_unitario':
                        prec = int(x.no_columna) - 1
                    if x.columna == 'cantidad':
                        qty = int(x.no_columna) - 1
                    if x.columna == 'importe':
                        imp = int(x.no_columna) - 1

            attachment = docto.attachment_id
            decoded_data = base64.b64decode(attachment.datas)
            workbook = openpyxl.load_workbook(filename=io.BytesIO(decoded_data), data_only=True)
            sheet = workbook.active
            registros = []
            contador = 1
            for row in sheet.iter_rows(values_only=True):
                if contador < inicio:
                    contador += 1
                    continue

                concepto = self.env['crm.concept.line'].search([('lead_id','=',record.id), ('concept_id.default_code','=',row[conc])])
                if concepto:
                    concept_id = concepto
                    continue

                if row[qty] == None:
                    continue
                if len(str(row[qty]).replace(' ', '')) == 0:
                    continue
                if not isinstance(row[imp], (int, float)):
                    continue

                registro = {'concept_id': concept_id.concept_id.id, 'col1': row[cod], 'col2': row[desc], 'col3': row[uni], 'col4': row[prec], 'col5': '', 
                    'col6': row[qty], 'col7': row[imp]}
                registros.append((0, 0, registro))
            record.write({'combo_ids': registros})


    def action_generar_combo(self):
        count = len(self.basico_ids.filtered(lambda u: not u.combo_ex))
        if count > 1:
            raise ValidationError('Faltan cargar los básicos')

        count = len(self.combo_ids.filtered(lambda u: not u.combo_ex))
        if count == 0:
            raise UserError('Las tarjetas de conceptos fueron generados correctamente')
        else:
            self.env.cr.execute('SELECT ccl.COL1, ccl.COL2 FROM crm_combo_line ccl WHERE ccl.LEAD_ID = ' + str(self.id) + ''' AND ccl.COMBO_EX IS FALSE 
                AND ccl.COL1 IS NOT NULL AND NOT EXISTS(SELECT * FROM product_template pt WHERE pt.DEFAULT_CODE = ccl.COL1) 
                AND NOT EXISTS(SELECT * FROM crm_basico_relacion cbr WHERE cbr.COL1 = ccl.COL1 and ccl.LEAD_ID = cbr.LEAD_ID) GROUP BY 1, 2 ORDER BY 1''')
            ex_basicos = self.env.cr.dictfetchall()
            mensaje = ''
            for x in ex_basicos:
                mensaje += x['col1'] + ' - ' + x['col2'] + '\n'
            if mensaje != '':
                raise ValidationError('Faltan de cargar los siguientes insumos del combo: %s' % mensaje)
            self.env.cr.execute('SELECT generar_combos(' + str(self.id)+ ', ' + str(self.env.user.id) + ')')
            basicos = self.env.cr.dictfetchall()
        

    def action_unlink_concept(self):
        for record in self:
            record.concept_ids.unlink()
            record.budget_ids.unlink()

    def _get_or_create_documents_folder(self, tipo):
        """ Crea/obtiene la estructura de carpetas en Documents
            CRM / <no_licitacion> / Tecnico o Economico """
        self.ensure_one()
        Document = self.env['documents.document']
        # Verificar si el módulo documents está disponible
        if 'documents.document' not in self.env:
            raise UserError(_('El módulo de Documentos no está disponible.\n Por favor, instale el módulo "documents" para usar esta funcionalidad.'))
        
        # 1) Carpeta raíz "CRM"
        root_folder = Document.search([('name', '=', 'CRM'), ('type', '=', 'folder'), ('folder_id', '=', False)], limit=1)
        if not root_folder:
            root_folder = Document.create({'name': 'CRM', 'type': 'folder', 'folder_id': False,})
        
        # 2) Carpeta de la licitación
        lic_name = self.no_licitacion or self.name or 'Sin nombre'
        lic_folder = Document.search([('name', '=', lic_name), ('type', '=', 'folder'), ('folder_id', '=', root_folder.id)], limit=1)
        if not lic_folder:
            lic_folder = Document.create({'name': lic_name, 'type': 'folder', 'folder_id': root_folder.id})
        
        # 3) Subcarpeta Técnico o Económico
        sub_folder = Document.search([('name', '=', tipo), ('type', '=', 'folder'), ('folder_id', '=', lic_folder.id)], limit=1)
        if not sub_folder:
            sub_folder = Document.create({'name': tipo, 'type': 'folder', 'folder_id': lic_folder.id,})
        
        return sub_folder


    def action_open_attachments(self):
        # Abre documentos en la carpeta específica (Técnico o Económico)
        self.ensure_one()
        # Determinar si es técnica o económica según el contexto
        tipo = self.env.context.get('folder_type', 'Tecnico')
        
        # Intentar usar Documents si está disponible
        if 'documents.document' in self.env:
            try:
                folder = self._get_or_create_documents_folder(tipo)
                # Buscar la acción de documents
                action = self.env.ref('documents.document_action').read()[0]
                
                # Filtrar por la carpeta específica
                action['domain'] = [('folder_id', '=', folder.id)]
                action['context'] = {'default_folder_id': folder.id, 'default_res_model': 'crm.lead', 'default_res_id': self.id, 
                    'searchpanel_default_folder_id': folder.id,}
                action['name'] = 'Documentos - %s - %s' % (self.no_licitacion or self.name, tipo)
                return action
                
            except Exception as e:
                _logger.error("Error usando Documents: %s" % str(e))
                raise UserError(_('Error al abrir documentos: %s') % str(e))
        else:
            raise UserError(_('El módulo de Documentos no está instalado.'))

    def action_view_documents(self):
        # Abre la carpeta de documentos de esta licitación
        self.ensure_one()
        if 'documents.document' in self.env:
            try:
                Document = self.env['documents.document']
                lic_name = self.no_licitacion or self.name or 'Sin nombre'
                # Buscar carpeta raíz CRM
                crm_folder = Document.search([('name', '=', 'CRM'), ('type', '=', 'folder'), ('folder_id', '=', False)], limit=1)
                if not crm_folder:
                    raise UserError(_('No se encontró la carpeta CRM.'))
                
                # Buscar carpeta de la licitación
                lic_folder = Document.search([('name', '=', lic_name), ('type', '=', 'folder'), ('folder_id', '=', crm_folder.id)], limit=1)
                if lic_folder:
                    # Abrir la carpeta de la licitación
                    action = self.env.ref('documents.document_action').read()[0]
                    action['context'] = {'default_folder_id': lic_folder.id, 'searchpanel_default_folder_id': lic_folder.id}
                    action['name'] = 'Documentos - %s' % lic_name
                    return action
                else:
                    # Si no existe la carpeta de la licitación, abrir CRM
                    action = self.env.ref('documents.document_action').read()[0]
                    action['context'] = {'default_folder_id': crm_folder.id, 'searchpanel_default_folder_id': crm_folder.id}
                    action['name'] = 'Documentos CRM'
                    return action
                
            except Exception as e:
                _logger.error("Error abriendo documentos: %s" % str(e))
                raise UserError(_('Error al abrir documentos: %s') % str(e))
        else:
            raise UserError(_('El módulo de Documentos no está instalado.'))


    def action_genera_ordenventa(self):
        self.ensure_one()
        if self.fecha_firma > self.fecha_limite_firma:
            raise UserError(_('La fecha de firma no puede ser mayor a la fecha límite de firma.'))

        if self.importe_contratado <= 0:
            raise UserError(_('Falta capturar el Importe Contratado.'))

        count = len(self.order_ids.filtered(lambda u: u.state != 'cancel'))
        if count != 0:
            raise ValidationError('Ya existe una orden de venta.')

        if not self.budget_ids or not self.concept_ids or not self.combo_ids or not self.basico_ids:
            raise ValidationError('Falta cargar información de los conceptos de trabajo y/o insumos')

        count = len(self.budget_ids.filtered(lambda u: not u.budget_id))
        if count != 0:
            raise ValidationError('Falta cargar las partidas')
            
        self.env.cr.execute('SELECT COUNT(*) num FROM crm_concept_line WHERE lead_id = ' + str(self.id) + 
            " AND CONCEPT_EX IS FALSE AND COL4 NOT IN ('',  NULL)")
        count = self.env.cr.dictfetchall()
        if count[0]['num'] > 0:
            raise ValidationError('Faltan cargar los conceptos de trabajo')

        count = len(self.combo_ids.filtered(lambda u: not u.combo_ex))
        if count > 1:
            raise ValidationError('Faltan cargar la matriz de insumos de los conceptos')

        count = len(self.basico_ids.filtered(lambda u: not u.combo_ex))
        if count > 1:
            raise ValidationError('Faltan cargar los básicos')

        oc_vals = self.get_work_default_values()
        oc_vals_2 = oc_vals[:]
        oc_new = self._create_work_async(oc_vals=oc_vals_2)
        return oc_new


    def get_work_default_values(self):
        sequence = self.env['ir.sequence'].next_by_code('sale.order')
        orders = []
        order_lines = []
        taxes = []
        fpos = self.env['account.fiscal.position'].with_company(self.company_id)._get_fiscal_position(self.partner_id)

        for rec in self.concept_ids.filtered(lambda u: u.concept_ex):
            taxes = rec.concept_id.taxes_id._filter_taxes_by_company(self.company_id)
            product_id = self.env['product.product'].search([('product_tmpl_id','=',rec.concept_id.id)])
            name = '[' + product_id.default_code + '] ' + product_id.name 

            lines = {'currency_id': product_id.currency_id.id, 'company_id': self.company_id.id, 'order_partner_id': self.partner_id.id, 
                'salesman_id': self.env.user.id, 'product_id': product_id.id, 'product_uom': product_id.uom_po_id.id, 'qty_delivered_method': 'timesheet',
                'invoice_status': 'to invoice', 'name': name, 'product_uom_qty': float(rec.col4.replace(',', '')), 
                'price_unit': float(rec.col5.replace(',', '')), 'tax_id': fpos.map_tax(taxes), 'state': 'sent', 'is_service': True, 
                'remaining_hours': float(rec.col4.replace(',', ''))}
            order_lines.append((0, 0, lines))

        tiene_anticipo = False
        if self.importe_anticipo != 0.00:
            tiene_anticipo = True

        values = {'company_id': self.company_id.id, 'partner_id': self.partner_id.id, 'partner_invoice_id': self.partner_id.id, 
            'partner_shipping_id': self.partner_id.id, 'currency_id': self.company_id.currency_id.id, 'user_id': self.env.uid, 'name': sequence, 
            'state': 'sent', 'origin': self.no_licitacion or self.name, 'invoice_status': 'to invoice', 'opportunity_id': self.id, 
            'tiene_anticipo': tiene_anticipo, 'anticipo_porcentaje': self.bases_anticipo_porcentaje, 'anticipo_importe': self.importe_anticipo, 
            'client_order_ref': self.no_licitacion or self.name, 'order_line': order_lines}
        orders.append(values)
        return orders

    def _create_work_async(self, oc_vals):
        oc_obj = self.env['sale.order']
        new_oc = oc_obj.create(oc_vals)
        return new_oc

    def action_set_lost(self, **kwargs):
        for lead in self:
            lead._ensure_stage_is_fallo()
        return super(CrmLead, self).action_set_lost(**kwargs)

    def write(self, vals):
        # Detectar cambio de "fallo_ganado" para disparar proceso automático
        change_fallo = 'fallo_ganado' in vals
        previous_fallo = {}
        if change_fallo:
            for lead in self:
                previous_fallo[lead.id] = lead.fallo_ganado

        # Cambios de documentos
        change_pt_docs = 'ptdcto_ids' in vals
        change_pe_docs = 'pedcto_ids' in vals

        res = super(CrmLead, self).write(vals)

        # Proceso automático de FALLO "Ganado"
        if change_fallo:
            for lead in self:
                if not previous_fallo.get(lead.id) and lead.fallo_ganado:
                    lead._send_fallo_notification(manual=False)

        # Sincronizar documentos de Propuesta Técnica
        if change_pt_docs:
            for lead in self:
                lead._sync_pt_doc_lines()

        if change_pe_docs:
            for lead in self:
                lead._sync_pe_doc_lines()
        return res


class crmInputsLine(models.Model):
    _name = 'crm.input.line'
    _description = 'Insumos'
    
    lead_id = fields.Many2one(comodel_name='crm.lead', string='Oportunidad', readonly=True)
    # T0054: Etiquetas de columnas ocultas (solo se muestran en la segunda fila de la vista)
    col1 = fields.Char(string='Código')
    col2 = fields.Char(string='Descripción')
    col3 = fields.Char(string='Unidad')
    col4 = fields.Char(string='Cantidad')
    col5 = fields.Char(string='Costo unitario')
    col6 = fields.Char(string='Importe')
    col7 = fields.Char(string='Columna 7')
    col8 = fields.Char(string='Columna 8')
    input_ex = fields.Boolean(string='Insumo cargado', default=False)
    input_id = fields.Many2one(comodel_name='product.template', string='Concepto de cobro')
    account_ex = fields.Boolean(string='Cuenta relacionada', default=False)
    tipinsumo_ex = fields.Boolean(string='Tipo de insumo relacionado', default=False)

class crmConceptLine(models.Model):
    _name = 'crm.concept.line'
    _description = 'Conceptos de trabajo'
    
    lead_id = fields.Many2one(comodel_name='crm.lead', string='Oportunidad', readonly=True)
    col1 = fields.Char(string='Código')
    col2 = fields.Char(string='Descripción')
    col3 = fields.Char(string='UDM')
    col4 = fields.Char(string='Cantidad')
    col5 = fields.Char(string='Precio Unitario')
    col6 = fields.Char(string='Importe')
    col7 = fields.Char(string='Columna 7')
    col8 = fields.Char(string='Columna 8')
    concept_ex = fields.Boolean(string='Concepto cargado', default=False)
    concept_id = fields.Many2one(comodel_name='product.template', string='Concepto de cobro')
    account_ex = fields.Boolean(string='Cuenta relacionada', default=False)
    combo_ex = fields.Boolean(string='Combo relacionado', default=False)

class crmBudgetLine(models.Model):
    _name = 'crm.budget.line'
    _description = 'Partidas presupuestarias'
    
    lead_id = fields.Many2one(comodel_name='crm.lead', string='Oportunidad', readonly=True)
    col1 = fields.Char(string='Codigo')
    col2 = fields.Char(string='Descripción')
    budget_id = fields.Many2one(comodel_name='product.budget.item', string='Partida')
    no_char = fields.Integer(string='No. de Caracteres')
    parent = fields.Integer(string='Padre')

class crmBasicoRelacion(models.Model):
    _name = 'crm.basico.relacion'
    _description = 'Relación de básicos'
    
    lead_id = fields.Many2one(comodel_name='crm.lead', string='Oportunidad', readonly=True)
    col1 = fields.Char(string='Código')
    col2 = fields.Char(string='Concepto')
    col3 = fields.Char(string='Unidad')
    col4 = fields.Char(string='Precio Unitario')
    combo_id = fields.Many2one(comodel_name='product.combo', string='Combo')

class crmBasicoLine(models.Model):
    _name = 'crm.basico.line'
    _description = 'Basicos de combos'
    
    lead_id = fields.Many2one(comodel_name='crm.lead', string='Oportunidad', readonly=True)
    basico = fields.Char(string='Básico')
    col1 = fields.Char(string='Código')
    col2 = fields.Char(string='Concepto')
    col3 = fields.Char(string='Unidad')
    col4 = fields.Char(string='Precio Unitario')
    col5 = fields.Char(string='Cantidad')
    col6 = fields.Char(string='Importe')
    combo_id = fields.Many2one(comodel_name='product.combo', string='Combo')
    relacion_id = fields.Many2one(comodel_name='crm.basico.relacion', string='Línea del combo')
    combo_ex = fields.Boolean(string='Combo relacionado', default=False)

class crmComboLine(models.Model):
    _name = 'crm.combo.line'
    _description = 'Combo por concepto'
    
    lead_id = fields.Many2one(comodel_name='crm.lead', string='Oportunidad', readonly=True)
    col1 = fields.Char(string='Código')
    col2 = fields.Char(string='Concepto')
    col3 = fields.Char(string='Unidad')
    col4 = fields.Char(string='Precio Unitario')
    col5 = fields.Char(string='Op.')
    col6 = fields.Char(string='Cantidad')
    col7 = fields.Char(string='Importe')
    concept_id = fields.Many2one(comodel_name='product.template', string='Concepto de cobro')
    combo_ex = fields.Boolean(string='Combo relacionado', default=False)

class CrmPropuestaTecnicaDoc(models.Model):
    _name = 'crm.propuesta.tecnica.doc'
    _description = 'Documentos Propuesta Técnica (CRM)'
    _rec_name = 'docto_id'

    lead_id = fields.Many2one('crm.lead', string='Oportunidad', required=True, ondelete='cascade',)
    docto_id = fields.Many2one('project.docsrequeridos', string='Docto', required=True, domain="[('model_id', '=', 'crm.lead'), ('etapa', '=', 'tecnica')]",)
    generado = fields.Boolean(string='Generado')

class CrmPropuestaTecnicaRevision(models.Model):
    _name = 'crm.propuesta.tecnica.revision'
    _description = 'Revisión de documentos Propuesta Técnica (CRM)'
    _rec_name = 'employee_id'
    _order = 'fecha_revision desc'
    
    lead_id = fields.Many2one('crm.lead', string='Oportunidad', required=True, ondelete='cascade')
    fecha_revision = fields.Datetime(string='Fecha de revisión', default=fields.Date.context_today, required=True)
    employee_id = fields.Many2one('hr.employee', string='Nombre', required=True, help='Persona que revisó la documentación.')
    revisado = fields.Boolean(string='Revisado', default=False)

class CrmPropuestaEconomicaDoc(models.Model):
    _name = 'crm.propuesta.economica.doc'
    _description = 'Documentos Propuesta Económica (CRM)'
    _rec_name = 'docto_id'

    lead_id = fields.Many2one('crm.lead', string='Oportunidad', required=True, ondelete='cascade',)
    docto_id = fields.Many2one('project.docsrequeridos', string='Docto', required=True, domain="[('model_id', '=', 'crm.lead'), ('etapa', '=', 'economica')]",)
    generado = fields.Boolean(string='Generado')

class CrmPropuestaEconomicaRevision(models.Model):
    _name = 'crm.propuesta.economica.revision'
    _description = 'Revisión de documentos Propuesta Económica (CRM)'
    _order = 'fecha_revision desc, id desc'
    
    lead_id = fields.Many2one('crm.lead', string='Oportunidad', required=True, ondelete='cascade')
    fecha_revision = fields.Datetime(string='Fecha de revisión', default=fields.Date.context_today, required=True)
    employee_ids = fields.Many2many('hr.employee', 'crm_pe_revision_employee_rel', 'revision_id', 'employee_id', string='Nombre', 
        help='Personas que revisarán la documentación.')
    revisado = fields.Boolean(string='Revisado', default=False)
    autorizado = fields.Boolean(string='Autorizado', default=False)
    activo = fields.Boolean(string='Activo', default=True, help='Indica si esta revisión está activa')

    @api.onchange('revisado')
    def _onchange_revisado(self):
        # Cuando se desmarca revisado, también desmarcar autorizado
        if not self.revisado and self.autorizado:
            self.autorizado = False
    
    @api.constrains('autorizado', 'revisado')
    def _check_autorizado(self):
        # No se puede autorizar sin haber revisado primero
        for record in self:
            if record.autorizado and not record.revisado:
                raise ValidationError(_('Debe marcar como "Revisado" antes de poder autorizar.'))

    @api.model_create_multi
    def create(self, vals_list):
        c = 0
        for vals in vals_list:
            lead_id = self.env['crm.lead'].search([('id', '=', vals.get('lead_id'))])
            revisiones_anteriores = self.env['crm.propuesta.economica.revision'].search([('lead_id', '=', vals.get('lead_id'))])
            if revisiones_anteriores:
                c = 1
                for rec in revisiones_anteriores:
                    rec.write({'activo': False})
                vals['autorizado'] = vals.get('autorizado')
            else:
                if vals.get('autorizado'):
                    vals['autorizado'] = False
                
            if c == 1 and vals['autorizado']:
                lead_id.pe_autorizado = True

        return super().create(vals_list)


    def write(self, values):
        if any(field in values for field in ['autorizado']):
            if 'lead_id' in values:
                lead = values.get('lead_id')
            else:
                lead = self.lead_id.id

            lead_id = self.env['crm.lead'].search([('id', '=', lead)])
            if values.get('autorizado'):
                lead_id.pe_autorizado = True
            else:
                lead_id.pe_autorizado = False

        return super().write(values)


class CrmJuntaLine(models.Model):
    _name = 'crm.junta.line'
    _description = 'Junta de Aclaración de Dudas'
    _order = 'fecha_hora asc'

    lead_id = fields.Many2one('crm.lead', string='Oportunidad', required=True, ondelete='cascade')
    fecha_hora = fields.Datetime(string='Fecha y hora', required=True,)
    descripcion = fields.Char(string='Descripción', size=250)
    lugar_reunion = fields.Char(string='Lugar de reunión', size=300)
    asistente_ids = fields.Many2many('hr.employee', 'crm_junta_line_employee_rel', 'junta_id', 'employee_id', string='Asistentes')
    notif_auto = fields.Boolean(string='Notif. Autom.', default=False)
    notif_manual = fields.Boolean(string='Notif. Manual', default=False)
    fecha_limite_preguntas = fields.Datetime(string='Fecha y hora límite para envío de preguntas')
    docto_preguntas = fields.Binary(string='Docto. Preguntas', attachment=True)
    docto_preguntas_name = fields.Char(string='Nombre docto. preguntas')
    acta_junta = fields.Binary(string='Acta de la junta', attachment=True)
    acta_junta_name = fields.Char(string='Nombre acta de la junta')

    def _get_asistente_emails(self):
        # Obtiene los correos de los asistentes asignados a esta junta
        self.ensure_one()
        emails = []
        for emp in self.asistente_ids:
            if emp.work_email:
                emails.append(emp.work_email)
            elif emp.user_id and emp.user_id.email:
                emails.append(emp.user_id.email)
        return emails

    def _send_junta_line_reminder(self, manual=False):
        """ Envía correo de recordatorio para esta línea de junta específica.
        
        Args:
            manual (bool): True si es notificación manual, False si es automática (cron) """
        self.ensure_one()
        template = self.env.ref('project_extra.mail_tmpl_junta_line_recordatorio', raise_if_not_found=False)

        if not template:
            raise UserError(_('No se encontró la plantilla de correo para recordatorio de Junta de Aclaración de Dudas. '
                            'Verifique que exista la plantilla "mail_tmpl_junta_line_recordatorio".'))
        if not self.fecha_hora:
            raise UserError(_('Debe capturar la fecha de la Junta de Aclaración de Dudas.'))
        if not self.asistente_ids:
            raise UserError(_('Debe asignar al menos un asistente para la Junta de Aclaración de Dudas.'))

        emails = self._get_asistente_emails()
        if not emails:
            raise UserError(_('Los asistentes asignados no tienen correo electrónico configurado.'))

        email_to = ', '.join(emails)
        try:
            template.send_mail(self.id, force_send=True, email_values={'email_to': email_to})
            
            if manual:
                self.write({'notif_manual': True})
            else:
                self.write({'notif_auto': True})
            
            self.lead_id.message_post(body=_("Se envió recordatorio de junta '%s' a: %s") % (self.descripcion or 'Sin descripción', email_to))
            _logger.info("Correo de junta enviado a: %s (Junta ID: %s)", email_to, self.id)
        except Exception as e:
            _logger.error("Error al enviar correo de junta (ID: %s): %s", self.id, str(e))
            self.lead_id.message_post(body=_("Error al enviar correo de junta: %s") % str(e))
            raise UserError(_('Error al enviar el correo: %s') % str(e))

    def action_send_notif_manual(self):
        # Botón para enviar notificación manual desde la línea de junta
        self.ensure_one()
        self._send_junta_line_reminder(manual=True)
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Notificación enviada'),
                'message': _('Se envió el correo de recordatorio correctamente a: %s') % ', '.join(self._get_asistente_emails()),
                'type': 'success',
                'sticky': False,}}

    def action_send_notif_auto(self):
        # Botón para enviar notificación automática manualmente (para pruebas o re-envío)
        self.ensure_one()
        self._send_junta_line_reminder(manual=False)
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Notificación enviada'),
                'message': _('Se envió el correo de recordatorio (automático) correctamente.'),
                'type': 'success',
                'sticky': False,}}


class CrmCircularLine(models.Model):
    _name = 'crm.circular.line'
    _description = 'Circular / Documento de dependencia'
    _order = 'fecha_hora asc'

    lead_id = fields.Many2one('crm.lead', string='Oportunidad', required=True, ondelete='cascade')
    fecha_hora = fields.Datetime(string='Fecha y hora', required=True)
    descripcion = fields.Char(string='Descripción', size=500)
    documento = fields.Binary(string='Documento', attachment=True)
    documento_name = fields.Char(string='Nombre del documento')
